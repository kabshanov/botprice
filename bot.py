# bot.py

import csv
import os
import random
import re
import tempfile
import openpyxl
import math
import itertools
import nest_asyncio
nest_asyncio.apply()

from telegram.ext import (
    ApplicationBuilder, CommandHandler, ContextTypes, CallbackContext,
    MessageHandler, filters, ConversationHandler, CallbackQueryHandler,
)
from google_drive_utils import API_TOKEN
from telegram import Update, BotCommand, InlineKeyboardButton, InlineKeyboardMarkup, Message, WebAppInfo
from telegram.constants import ChatAction, ParseMode
from datetime import datetime, timezone
from typing import Tuple, List, Dict, Any, Set
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

from shared_data import (USER_DATA, PRODUCT_LIBRARY, TEXT_TO_FLAG, MAX_TELEGRAM_TEXT, LIST_BRAND_GROUP,
    currency_api_cbr, exchange_api)
from db_utils import init_db, clear_user_data
from utils import message_handler
from sorting_rules import get_sort_key
from currency_api import BinanceAPI
from ai_assistant import register_ai_assistant


import logging
logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.WARNING, #⚠️⚠️
    datefmt='%H:%M'
)
# Отключение избыточных логов библиотеки httpx
logging.getLogger("httpx").setLevel(logging.WARNING)

# ------------------ Утилита: pseudo_user_data -------------------
def load_pseudo_user_data(user_id: int) -> dict:
    """
    Вспомогательная функция:
    Возвращает словарь {"products": [...]} из базы,
    чтобы можно было без переписывания кода
    использовать build_final_product_list({ "products": [...] }).
    """
    rows = get_all_products(user_id)
    return {"products": rows}

#-----------------------дополнительный обработчики ввыводе-----------------------------

def remove_commented_products(final_product_list):
    """
    Убирает из final_product_list все товары, у которых есть непустой comment.
    """
    for mg in list(final_product_list.keys()):
        product_dict = final_product_list[mg]
        for product_name in list(product_dict.keys()):
            entries = product_dict[product_name]
            # Оставляем только те, где comment=''
            filtered = [e for e in entries if not e["comment"].strip()]
            if filtered:
                product_dict[product_name] = filtered
            else:
                # Пусто — удаляем
                del product_dict[product_name]
        # Если model_group пуст
        if not product_dict:
            del final_product_list[mg]

def escape_html(text: str) -> str:
    """
    Экранирует спецсимволы для корректного использования в HTML-разметке Telegram.
    """
    return (text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace('"', "&quot;"))

#-----------------------Сброс всех состояний-----------------------------
async def restart_all(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    # 1) Очищаем user_data
    context.user_data.clear()

    # 2) Отправляем ответ
    await update.message.reply_text("Все состояния сброшены! Можете начинать заново.")

    # 3) Возвращаем ConversationHandler.END
    #    чтобы принудительно выйти из любого ConversationHandler
    return ConversationHandler.END

#-----------------------my_price_list-----------------------------
from db_utils import (
    get_all_products,
    get_user_settings,
    save_user_settings,
    clear_user_settings,
    # ...
)
...

# Вспомогательные функции для ConversationHandler

def reset_error_counter(context: ContextTypes.DEFAULT_TYPE, key: str) :
    context.user_data[key] = 0

def increment_error_counter(context: ContextTypes.DEFAULT_TYPE, key: str) :
    context.user_data[key] = context.user_data.get(key, 0) + 1

def get_error_counter(context: ContextTypes.DEFAULT_TYPE, key: str) :
    return context.user_data.get(key, 0)

async def send_typing_action(context: ContextTypes.DEFAULT_TYPE, chat_id: int) :
    await context.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
    await asyncio.sleep(1)  # Имитация "печатающего бота"

# ConversationHandler states
CHOOSE_COLUMNS = 1
ENTER_COLUMNS_NAMES = 2
ENTER_PRICES_GRADATION = 3
CHOOSE_FORMAT_LIST = 4

# Сообщения об ошибках для этапа выбора количества столбцов
ERROR_MESSAGES_COLUMNS = [
    "🙈 Упс! Это не то число. Пожалуйста, введите от 1 до 5. Попробуем снова? 🔄",
    "🧐 Что-то пошло не так. Давайте ещё раз: от 1 до 5, не больше и не меньше. 😊",
    "🤔 Хм, кажется, я ждал число от 1 до 5. Давайте попробуем ещё раз. ⏳",
    "🔢 Похоже, вы ввели что-то не то. Я принимаю только числа от 1 до 5! Дерзайте! 💪",
    "❗ Ой-ой, такое число мне не подходит. Нужно что-то между 1 и 5. Попробуем снова? 🚀",
    "🤷‍♀️ Ой, кажется, я запутался... Я жду число от 1 до 5. Попробуем ещё раз? 🔄",
    "🔍 Кажется, что-то пошло не так! Проверьте ввод: мне нужно число от 1 до 5. Верю, у вас получится! 💪",
    "🧮 Мой калькулятор не понимает это число... Введите число от 1 до 5, и мы продолжим. 🚀",
    "😅 Я пока ещё учусь считать... Помогите мне! Введите число между 1 и 5, пожалуйста. 🧡",
    "⚠️ Похоже, вы вышли за рамки! Числа от 1 до 5 – это мой предел. Попробуйте ещё раз! 🔢"
]

FINAL_ERROR_COLUMNS = [
    "😅 Похоже, что-то пошло не так! Давайте начнём заново? Используйте команду /my_price_list.",
    "🤷‍♂️ Видимо, это не ваш день с числами! Попробуем с начала? Наберите /my_price_list.",
    "🎭 Математика бывает сложной! Давайте вернёмся к началу. /my_price_list поможет."
]

# Сообщения об ошибках для этапа ввода названий столбцов
ERROR_MESSAGES_COLUMNS_LENGTH = [
    "⏳ Ой, одно из названий столбцов слишком длинное (максимум 30 символов). Давайте немного сократим его и попробуем ещё раз. 😉",
    "😅 Похоже, вы разогнались с названием. Попробуйте уложиться в 30 символов – лаконичность рулит! 🚀",
    "🧐 Упс! Название слишком длинное... Мне подойдёт что-то покороче – максимум 30 символов. Попробуем снова? ✂️",
    "🙈 Название слишком длинное! Помогите мне справиться – введите название короче 30 символов. 😊",
    "📏 Ого! Кажется, название столбца растянулось слишком сильно. Максимум 30 символов – давайте укоротим и попробуем снова. 💡",
    "✂️ Ай-ай! Название слишком длинное... Давайте обрежем лишнее и уложимся в 30 символов. Коротко и понятно! 🚀",
    "🔍 Упс, одно из названий выходит за рамки! Помним: максимум 30 символов. Давайте вернёмся и поправим. 📋",
    "🖊️ Такое длинное название просто не влезет в мой файл! Попробуйте уложиться в 30 символов, и всё будет отлично. 🎉"
]

FINAL_ERROR_COLUMNS_LENGTH = [
    "😅 Похоже, вы любите длинные названия! Давайте начнём сначала. Введите команду /my_price_list. 📝",
    "🤷‍♂️ Кажется, мы слегка запутались в буквах... Попробуем с самого начала? Наберите /my_price_list. 🔄",
    "🎭 Пора сделать передышку! Уберите лишнее и попробуйте заново с командой /my_price_list. Я верю в вас! 💪",
    "🛠 Что-то идёт не так... Может, сократим и начнём сначала? Введите /my_price_list для перезапуска. 🚀",
    "🧹 Давайте расчистим дорогу для новых идей! Начинаем заново с командой /my_price_list. ⏳"
]

# Сообщения об ошибках для этапа градации цен
ERROR_MESSAGES_PRICES_FORMAT = [
    "🔄 Ой! Похоже, формат ввода немного не соответствует примеру. Пожалуйста, проверьте и попробуйте снова. 😊",
    "📝 Кажется, в вашем вводе есть неточности. Убедитесь, что вы следуете примеру, и повторите попытку. 🚀",
    "❗ Внимание! Формат данных не распознан. Пожалуйста, введите градацию цен согласно образцу. 💡",
    "🤔 Хм, что-то не так. Введите данные в правильном формате. 📋",
    "🔢 Похоже, числа или диапазоны указаны неверно. Проверьте формат и попробуйте ещё раз. 🧮",
    "🛠 Упс! Возникла ошибка в формате ввода. Давайте сверимся с примером и повторим попытку. 📝",
    "📊 Данные не соответствуют формату. Используйте пример в качестве ориентира. 😊",
    "🙈 Я не смог обработать ваш ввод. Убедитесь, что вы следуете структуре примера, и попробуйте снова. 🔄",
    "🚧 Ошибка! Формат градации цен некорректен. Давайте проверим ещё раз по примеру и повторим. 💪",
    "🔍 Не удаётся распознать введённые данные. Пожалуйста, внимательно следуйте примеру и попробуйте снова. 📖"
]

FINAL_ERROR_PRICES_FORMAT = [
    "😅 Кажется, это непросто! Давайте начнём сначала. Введите команду /my_price_list, чтобы попробовать ещё раз. 🌟",
    "🤷‍♂️ Не беда, иногда всем нужна перезагрузка. Наберите /my_price_list, и мы начнём заново. 🚀",
    "🎯 Давайте попробуем с начала! Используйте команду /my_price_list, чтобы перезапустить процесс. Уверен, в этот раз всё получится! 💪"
]

# Функции этапов ConversationHandler

async def start_my_price_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # 1. Сначала получаем текущие настройки из БД
    db_settings = get_user_settings(user_id)
    if db_settings:
        # Запишем в context.user_data, чтобы продолжить
        context.user_data["num_columns"] = db_settings["num_columns"]
        context.user_data["columns_names"] = db_settings["columns_names"]
        context.user_data["prices_gradation"] = db_settings["prices_gradation"]
        context.user_data["output_format"] = db_settings["output_format"]

    # Проверяем, есть ли у пользователя товары
    products = get_all_products(user_id)
    if not products:
        await update.message.reply_text(
            "⚠️ Ваш список товаров пуст. Пожалуйста, добавьте товары перед созданием прайс-листа."
        )
        return ConversationHandler.END

    # Если все нужные настройки уже есть => просим выбрать формат
    if (
        "columns_names" in context.user_data and
        "prices_gradation" in context.user_data and
        "output_format" in context.user_data
    ):
        await choose_output_format(update, context)  # сразу на выбор
        return CHOOSE_FORMAT_LIST

    # Сброс ошибок
    reset_error_counter(context, "columns_error")
    reset_error_counter(context, "columns_length_error")
    reset_error_counter(context, "prices_format_error")

    context.user_data["my_price_list_state"] = CHOOSE_COLUMNS

    await send_typing_action(context, update.effective_chat.id)
    await update.message.reply_text(
        "🚀 Отлично! Давайте создадим ваш прайс-лист. Сколько столбцов вам понадобится – от 1 до 5? 😊"
    )
    return CHOOSE_COLUMNS

async def choose_columns(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if text.isdigit():
        num = int(text)
        if 1 <= num <= 5:
            reset_error_counter(context, "columns_error")
            context.user_data["num_columns"] = num
            # Сохраняем в БД
            save_user_settings(
                user_id,
                num_columns=num  # остальное пока None
            )

            if num == 1:
                msg = (
                    "📝 Введите название столбца:\n"
                    "Пример:\nНаименование столбца"
                )
            else:
                msg = f"📝 Введите названия {num} столбцов (каждое с новой строки). Пример:\n"
                for i in range(1, num + 1):
                    msg += f"Наименование столбца {i}\n"

            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(msg)
            return ENTER_COLUMNS_NAMES
        else:
            increment_error_counter(context, "columns_error")
            err_count = get_error_counter(context, "columns_error")
            if err_count < 5:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(ERROR_MESSAGES_COLUMNS))
                return CHOOSE_COLUMNS
            else:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(FINAL_ERROR_COLUMNS))
                return ConversationHandler.END
    else:
        increment_error_counter(context, "columns_error")
        err_count = get_error_counter(context, "columns_error")
        if err_count < 5:
            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(random.choice(ERROR_MESSAGES_COLUMNS))
            return CHOOSE_COLUMNS
        else:
            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(random.choice(FINAL_ERROR_COLUMNS))
            return ConversationHandler.END

async def enter_columns_names(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    num = context.user_data.get("num_columns")
    text = update.message.text.strip()
    lines = text.split("\n")

    if len(lines) != num:
        increment_error_counter(context, "columns_length_error")
        err_count = get_error_counter(context, "columns_length_error")
        if err_count < 5:
            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(
                "⏳ Ой, количество строк не совпадает с количеством столбцов. Попробуем ещё раз."
            )
            return ENTER_COLUMNS_NAMES
        else:
            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(random.choice(FINAL_ERROR_COLUMNS_LENGTH))
            return ConversationHandler.END

    # Проверка длины названий
    for col_name in lines:
        if len(col_name) > 30:
            increment_error_counter(context, "columns_length_error")
            err_count = get_error_counter(context, "columns_length_error")
            if err_count < 5:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(ERROR_MESSAGES_COLUMNS_LENGTH))
                return ENTER_COLUMNS_NAMES
            else:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(FINAL_ERROR_COLUMNS_LENGTH))
                return ConversationHandler.END

    reset_error_counter(context, "columns_length_error")
    context.user_data["columns_names"] = lines
    # Сохраняем в БД
    save_user_settings(
        user_id,
        columns_names=lines  # num_columns уже есть в БД
    )

    # Переходим к этапу градации цен
    num_columns = num
    # Создаём пример с разными наценками, чтобы показать разнообразие
    example_lines = [
        "0-10000 " + " ".join([f"{1000 * (i + 1)}" for i in range(num_columns)]),
        "10001-50000 " + " ".join([f"{1500 * (i + 1)}" for i in range(num_columns)]),
        "50001-100000 " + " ".join([f"{2000 * (i + 1)}" for i in range(num_columns)]),
        "100001-200000 " + " ".join([f"{7500 * (i + 1)}" for i in range(num_columns)])
    ]
    example = (
            f"Отлично! Теперь приступим к настройке градации цен для вашего прайс-листа.\n\n"
            f"📋 Пример ввода для {num_columns} дополнительных столбцов:\n\n"
            + "\n".join(example_lines) +
            "\n\nОбъяснение ввода:\n"
            "Первый диапазон цен начинается с 0 до N.\n"
            "Далее идут последовательные диапазоны цен – каждый диапазон должен логически следовать за предыдущим, без пропусков чисел.\n"
            "Пример: 10001-50000 следует за 10000.\n"
            "Последняя строка – это диапазон от N до N, который определяет верхний предел цен и заканчивает градацию.\n\n"
            f"🗒️ Важно:\n\n"
            f"Количество значений после каждого диапазона должно соответствовать числу дополнительных столбцов, которые вы указали ранее ({num_columns}).\n"
            f"В нашем примере их {num_columns}: {', '.join(context.user_data['columns_names'])}.\n"
            "Пожалуйста, введите ваши значения градации цен в соответствии с примером."
    )

    await send_typing_action(context, update.effective_chat.id)
    await update.message.reply_text(example)
    return ENTER_PRICES_GRADATION

async def enter_prices_gradation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    num = context.user_data.get("num_columns")
    text = update.message.text.strip()
    lines = text.split("\n")
    gradations = []

    for line in lines:
        parts = line.split()
        if len(parts) != 1 + num:
            increment_error_counter(context, "prices_format_error")
            err_count = get_error_counter(context, "prices_format_error")
            if err_count < 5:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(ERROR_MESSAGES_PRICES_FORMAT))
                return ENTER_PRICES_GRADATION
            else:
                await send_typing_action(context, update.effective_chat.id)
                await update.message.reply_text(random.choice(FINAL_ERROR_PRICES_FORMAT))
                return ConversationHandler.END

        # Проверка формата диапазона и чисел
        range_part = parts[0]
        increments = parts[1:]
        if '-' in range_part:
            range_bounds = range_part.split('-')
            if len(range_bounds) != 2 or not all(rb.isdigit() for rb in range_bounds):
                increment_error_counter(context, "prices_format_error")
                err_count = get_error_counter(context, "prices_format_error")
                if err_count < 5:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(ERROR_MESSAGES_PRICES_FORMAT))
                    return ENTER_PRICES_GRADATION
                else:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(FINAL_ERROR_PRICES_FORMAT))
                    return ConversationHandler.END
        else:
            if not range_part.isdigit():
                increment_error_counter(context, "prices_format_error")
                err_count = get_error_counter(context, "prices_format_error")
                if err_count < 5:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(ERROR_MESSAGES_PRICES_FORMAT))
                    return ENTER_PRICES_GRADATION
                else:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(FINAL_ERROR_PRICES_FORMAT))
                    return ConversationHandler.END

        # Проверка наценок
        for inc in increments:
            if not inc.isdigit():
                increment_error_counter(context, "prices_format_error")
                err_count = get_error_counter(context, "prices_format_error")
                if err_count < 5:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(ERROR_MESSAGES_PRICES_FORMAT))
                    return ENTER_PRICES_GRADATION
                else:
                    await send_typing_action(context, update.effective_chat.id)
                    await update.message.reply_text(random.choice(FINAL_ERROR_PRICES_FORMAT))
                    return ConversationHandler.END

        # Добавляем диапазон и наценки
        if '-' in range_part:
            start, end = map(int, range_part.split('-'))
            gradations.append((start, end, list(map(int, increments))))
        else:
            val = int(range_part)
            gradations.append((val, None, list(map(int, increments))))

    # Проверяем логичность последовательности диапазонов
    sorted_gradations = sorted(gradations, key=lambda x: x[0])
    previous_end = 0
    for i, (start, end, _) in enumerate(sorted_gradations):
        if start != previous_end + 1 and previous_end != 0:
            await send_typing_action(context, update.effective_chat.id)
            await update.message.reply_text(
                "❗ Диапазоны цен должны логически следовать друг за другом без пропусков чисел. Попробуйте снова."
            )
            return ENTER_PRICES_GRADATION
        if end:
            previous_end = end
        else:
            previous_end = start + 1  # Для последнего диапазона
    # Все проверки пройдены

    # Если все проверки пройдены
    reset_error_counter(context, "prices_format_error")
    context.user_data["prices_gradation"] = gradations
    # Сохраняем в БД
    save_user_settings(
        user_id,
        prices_gradation=gradations
    )

    # >>> СООБЩАЕМ ПОЛЬЗОВАТЕЛЮ, ЧТО НАСТРОЙКИ СОХРАНЕНЫ:
    await update.message.reply_text(
        "✅ Настройки прайс-листа сохранены!\n"
        "При следующем вызове команды будет использован этот формат.\n"
        "Чтобы сбросить настройки, отправьте /restart_price_list."
    )

    await choose_output_format(update, context)
    return CHOOSE_FORMAT_LIST

async def cancel_my_price_list(update: Update, context: ContextTypes.DEFAULT_TYPE) :
    user_id = update.effective_user.id
    if user_id in USER_DATA :
        # Удаляем данные, связанные с my_price_list
        USER_DATA[user_id].pop("my_price_list_state", None)
        USER_DATA[user_id].pop("num_columns", None)
        USER_DATA[user_id].pop("columns_names", None)
        USER_DATA[user_id].pop("prices_gradation", None)
    await update.message.reply_text(
        "Процесс создания прайс-листа отменён. Используйте /my_price_list для нового запуска.")
    return ConversationHandler.END

async def choose_output_format(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("CSV", callback_data="CSV"),
            InlineKeyboardButton("Excel", callback_data="Excel"),
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("📂 Выберите формат для экспорта прайс-листа:", reply_markup=reply_markup)
    return CHOOSE_FORMAT_LIST

async def choose_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    choice = query.data
    if choice not in ["CSV", "Excel"]:
        await query.edit_message_text("❌ Неверный выбор формата. Пожалуйста, выберите CSV или Excel.")
        return CHOOSE_FORMAT_LIST

    # Сохраняем выбранный формат
    context.user_data["output_format"] = choice

    # Удаляем кнопки из сообщения
    await query.edit_message_reply_markup(reply_markup=None)

    # Генерируем прайс-лист в выбранном формате
    if choice == "CSV":
        await generate_price_list_csv(update, context)
    elif choice == "Excel":
        await generate_price_list_excel(update, context)

    return ConversationHandler.END

def get_my_price_list_conversation_handler():
    return ConversationHandler(
        entry_points=[CommandHandler("my_price_list", start_my_price_list)],
        states={
            CHOOSE_COLUMNS: [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_columns)],
            ENTER_COLUMNS_NAMES: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_columns_names)],
            ENTER_PRICES_GRADATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, enter_prices_gradation)],
            CHOOSE_FORMAT_LIST: [CallbackQueryHandler(choose_format_callback, pattern="^(CSV|Excel)$")],
        },
        fallbacks=[CommandHandler("cancel", cancel_my_price_list)],
        name="my_price_list_conversation",
        persistent=False
    )

def find_brand_for_model_group(model_group: str) -> str:
    """
    Если всё-таки надо определить бренд «на лету»,
    но мы хотим явно использовать структуру PRODUCT_LIBRARY
    и очередность, можем не делать этот метод, если уже
    знаем brand при итерации.
    """
    for brand, brand_data in PRODUCT_LIBRARY.items():
        if model_group in brand_data:
            return brand
    return "Unknown"

async def generate_price_list_csv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_name = update.effective_user.username or f"User_{user_id}"

    columns_names = context.user_data.get("columns_names", [])
    prices_gradation = context.user_data.get("prices_gradation", [])

    # Вместо:
    #   products = USER_DATA[user_id].get("products", [])
    #   if not products: ...
    # теперь:
    products = get_all_products(user_id)
    if not columns_names or not prices_gradation:
        await update.effective_message.reply_text("Настройки прайс-листа не завершены. Пожалуйста, используйте команду /my_price_list для настройки.")
        return

    if not products:
        await update.effective_message.reply_text(
            "Ваш список товаров пуст. Отправьте сообщения перед использованием команды."
        )
        logging.info(f"Пустой список товаров для пользователя {user_name} ({user_id})")
        return

    logging.info(f"[generate_price_list_csv] Начало формирования праййс-листа CSV для {user_name} ({user_id})")

    # ЗАГРУЖАЕМ "как будто" user_data
    pseudo_data = load_pseudo_user_data(user_id)
    # теперь вызываем build_final_product_list(pseudo_data)
    best_offers = build_final_product_list(pseudo_data)

    # Дальше логика не меняется:
    # формируем flat, фильтруем товары с комментариями и т.д.
    # ...
    # (код из вашего примера оставляем, только заменили products -> best_offers)
    # --------------------------------------------------------------------------------
    flat_best_offers = []
    for model_group, products_dict in best_offers.items():
        for product_name, offers in products_dict.items():
            for offer in offers:
                flat_best_offers.append({
                    "model_group": model_group,
                    "product_name": product_name,
                    "price": offer["price"],
                    "country": offer["country"],
                    "supplier": offer["suppliers"],
                    "comment": offer["comment"]
                })

    filtered_offers = [o for o in flat_best_offers if not o.get("comment", "").strip()]
    if not filtered_offers:
        await update.effective_message.reply_text(
            "⚠️ После фильтрации нет товаров без комментариев для формирования прайс-листа."
        )
        logging.info(f"Нет товаров без комментов у {user_name} ({user_id})")
        return

    def get_increments_for_price(price):
        for start, end, incs in sorted(prices_gradation, key=lambda x: x[0]):
            if end is None:
                if price >= start:
                    return incs
            else:
                if start <= price <= end:
                    return incs
        return [0] * len(columns_names)

    grouped = defaultdict(list)
    for item in filtered_offers:
        grouped[item["model_group"]].append(item)

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"my_price_list_{current_date}.csv"

    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = ["Наименование", "Страна"] + columns_names
            writer.writerow(header)

            # Соблюдаем иерархию PRODUCT_LIBRARY
            for brand, model_groups in PRODUCT_LIBRARY.items():
                for mg in model_groups.keys():
                    if mg not in grouped:
                        continue

                    writer.writerow([mg, "Страна"] + columns_names)

                    items = grouped[mg]

                    # Добавляем вторым элементом x["price"]:
                    sorted_items = sorted(
                        items,
                        key=lambda x : (
                            get_sort_key(brand, mg, x["product_name"]),
                            x["price"]  # <- сортируем по возрастанию цены, если product_name совпадает
                        )
                    )

                    for row_item in sorted_items :
                        price_val = row_item["price"]
                        incs = get_increments_for_price(price_val)
                        new_prices = [price_val + inc for inc in incs]

                        writer.writerow([
                                            row_item["product_name"],
                                            row_item["country"]
                                        ] + new_prices)

                    # Пустая строка после группы
                    writer.writerow([])

        await send_typing_action(context, update.effective_chat.id)
        try:
            with open(filename, "rb") as f:
                await update.effective_message.reply_document(
                    document=f,
                    filename=filename,
                    caption="Ваш прайс-лист (CSV) готов!"
                )
        except Exception as e:
            logging.error(f"Ошибка при отправке CSV: {e}")
            await update.effective_message.reply_text("Произошла ошибка при отправке CSV.")
    except Exception as e:
        logging.error(f"Ошибка формирования CSV: {e}")
        await update.effective_message.reply_text("Произошла ошибка при формировании CSV файла.")
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления файла {filename}: {e}")

    logging.info(f"[generate_price_list_csv] Завершено для {user_name} ({user_id})")

async def generate_price_list_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_name = update.effective_user.username or f"User_{user_id}"

    columns_names = context.user_data.get("columns_names", [])
    prices_gradation = context.user_data.get("prices_gradation", [])

    products = get_all_products(user_id)

    if not columns_names or not prices_gradation:
        await update.effective_message.reply_text(
            "Настройки прайс-листа не завершены. Пожалуйста, используйте команду /my_price_list ..."
        )
        return

    if not products:
        await update.effective_message.reply_text("Ваш список товаров пуст...")
        logging.info(f"Пустой список для {user_name} ({user_id})")
        return

    logging.info(f"[generate_price_list_excel] Старт формирования Excel для {user_name} ({user_id})")

    # псевдо-данные
    pseudo_data = load_pseudo_user_data(user_id)
    best_offers = build_final_product_list(pseudo_data)

    # далее код не меняется
    flat_best_offers = []
    for mg, prod_dict in best_offers.items():
        for product_name, offers in prod_dict.items():
            for offer in offers:
                flat_best_offers.append({
                    "model_group": mg,
                    "product_name": product_name,
                    "price": offer["price"],
                    "country": offer["country"],
                    "supplier": offer["suppliers"],
                    "comment": offer["comment"]
                })

    filtered_offers = [o for o in flat_best_offers if not o.get("comment", "").strip()]
    if not filtered_offers:
        await update.effective_message.reply_text(
            "⚠️ После фильтрации нет товаров без комментариев ..."
        )
        return

    def get_increments_for_price(price):
        for start, end, incs in sorted(prices_gradation, key=lambda x: x[0]):
            if end is None:
                if price >= start:
                    return incs
            else:
                if start <= price <= end:
                    return incs
        return [0] * len(columns_names)

    def sort_products(items) :
        return sorted(
            items,
            key=lambda x : (
                get_sort_key(
                    find_brand_for_model_group(x["model_group"]),
                    x["model_group"],
                    x["product_name"]
                ),
                x["price"]  # вторичный критерий: цена по возрастанию
            )
        )

    grouped = defaultdict(list)
    for row_item in filtered_offers:
        grouped[row_item["model_group"]].append(row_item)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    header = ["Наименование", "Страна"] + columns_names
    ws.append(header)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for brand, mg_dict in PRODUCT_LIBRARY.items():
        for mg in mg_dict.keys():
            if mg not in grouped:
                continue
            ws.append([mg, "Страна"] + columns_names)
            sorted_items = sort_products(grouped[mg])
            for row_item in sorted_items:
                incs = get_increments_for_price(row_item["price"])
                new_prices = [row_item["price"] + inc for inc in incs]
                ws.append([
                    row_item["product_name"],
                    row_item["country"]
                ] + new_prices)
            ws.append([])

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"my_price_list_{current_date}.xlsx"

    try:
        wb.save(filename)
    except Exception as e:
        logging.error(f"Ошибка при сохранении Excel: {e}")
        await update.effective_message.reply_text("Произошла ошибка при формировании Excel файла.")
        return

    await send_typing_action(context, update.effective_chat.id)
    try:
        with open(filename, "rb") as f:
            await update.effective_message.reply_document(
                document=f,
                filename=filename,
                caption="Ваш прайс-лист Excel готов!"
            )
    except Exception as e:
        logging.error(f"Ошибка отправки Excel: {e}")
        await update.effective_message.reply_text("Произошла ошибка при отправке Excel.")
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления {filename}: {e}")

    logging.info(f"[generate_price_list_excel] Завершено для {user_name} ({user_id})")

async def restart_price_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Очищаем настройки из БД и из context
    clear_user_settings(user_id)

    keys_to_remove = ["my_price_list_state", "num_columns", "columns_names", "prices_gradation", "output_format"]
    for key in keys_to_remove:
        context.user_data.pop(key, None)

    await update.message.reply_text(
        "🛠 Настройки прайс-листа сброшены. При следующем использовании команды /my_price_list вы начнёте заново."
    )
    return ConversationHandler.END


# ------------------------------ ФУНКЦИИ ДЛЯ ФОРМИРОВАНИЯ BEST------------------------------

def build_final_product_list(user_data: dict) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Вытаскивает товары из user_data, группирует их по model_group, product_name и country,
    применяет сценарии 1-5 для выбора лучших предложений, и возвращает final_product_list.

    Новые доработки:
      1. Если один и тот же товар (model_group, product_name) имеет одинаковую price,
         и среди них есть вариант(ы) со страной и без – при равной цене
         сохраняем только вариант(ы) со страной, удаляем без-страны.
      2. Если у одного товара одинаковая price, одна и та же страна, но разные supplier —
         оставляем только единственного "лучшего" по алфавиту supplier.
      3. Аналогично, если одинаковые товары с одинаковой ценой без страны —
         выбираем лучший по поставщику.
      4. Дополнительный шаг: если для одного (model_group, product_name) имеются записи с известной страной
         и с пустой (""), то для всех записей с пустой страной оставляем только те, у которых цена ниже или равна
         минимальной цене среди записей с известной страной.
    """
    products_data = user_data.get("products", [])
    grouped_products: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)

    # Группировка товаров по (model_group, product_name, country)
    for message_data in products_data :
        line = message_data.get("line", "")
        country = message_data.get("country", "")
        product_name = message_data.get("product_name")
        model_group = message_data.get("model_group")
        price = message_data.get("price")
        supplier = message_data.get("supplier") or ""  # Исправлено здесь
        comment = message_data.get("comment", "")

        if not product_name or not model_group or price is None :
            logging.info(f"Пропущена строка из-за отсутствия данных: {line}")
            continue

        group_key = (model_group, product_name, country)
        grouped_products[group_key].append({
            "price" : price,
            "supplier" : supplier,
            "comment" : comment
        })

        if not product_name or not model_group or price is None:
            logging.info(f"Пропущена строка из-за отсутствия данных: {line}")
            continue

        group_key = (model_group, product_name, country)
        grouped_products[group_key].append({
            "price": price,
            "supplier": supplier,
            "comment": comment
        })

    final_product_data: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}

    # Шаг 1: Удаление полностью идентичных предложений (одинаковые price / supplier / comment)
    for key, raw_products in grouped_products.items():
        unique_offers = []
        seen_offers = set()
        for p in raw_products:
            offer_tuple = (p["price"], p["supplier"], p["comment"])
            if offer_tuple not in seen_offers:
                unique_offers.append(p)
                seen_offers.add(offer_tuple)
        grouped_products[key] = unique_offers

    # Шаг 2: Группировка по (model_group, product_name) независимо от country
    mg_pn_map: Dict[Tuple[str, str], List[Tuple[str, int, str, str]]] = defaultdict(list)
    for (mg, pn, ctr), offers in grouped_products.items():
        for p in offers:
            mg_pn_map[(mg, pn)].append((ctr, p["price"], p["supplier"], p["comment"]))

    # Обработка внутри каждой пары (model_group, product_name)
    new_grouped_products: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for (mg, pn), items in mg_pn_map.items():
        # Сгруппируем записи по цене
        price_map: Dict[int, List[Tuple[str, str, str]]] = defaultdict(list)
        for (ctr, pr, sup, cmt) in items:
            price_map[pr].append((ctr, sup, cmt))
        # Для каждой цены
        for pr_value, sublist in price_map.items():
            # Если есть хотя бы один вариант с известной страной, оставляем только их
            has_known_country = any(ctr != "" for (ctr, sup, cmt) in sublist)
            if has_known_country:
                sublist = [(ctr, sup, cmt) for (ctr, sup, cmt) in sublist if ctr != ""]
            # Для каждой страны (или пустой) с одинаковой ценой выбираем один вариант по supplier
            country_map: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
            for (ctr, sup, cmt) in sublist:
                country_map[ctr].append((sup, cmt))
            for ctr, sup_list in country_map.items():
                if len(sup_list) > 1:
                    sup_list_sorted = sorted(sup_list, key=lambda x: x[0].lower())
                    best_sup, best_cmt = sup_list_sorted[0]
                    new_grouped_products[(mg, pn, ctr)].append({
                        "price": pr_value,
                        "supplier": best_sup,
                        "comment": best_cmt
                    })
                else:
                    single_sup, single_cmt = sup_list[0]
                    new_grouped_products[(mg, pn, ctr)].append({
                        "price": pr_value,
                        "supplier": single_sup,
                        "comment": single_cmt
                    })

    # Шаг 3: Дополнительная фильтрация по стране:
    # Для каждой пары (model_group, product_name), если существуют записи с известной страной,
    # удаляем из группы все записи с пустой страной, если их цена выше минимальной цены среди известных.
    final_group: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for (mg, pn, ctr), recs in new_grouped_products.items():
        final_group[(mg, pn)].extend([dict(record, country=ctr) for record in recs])
    adjusted_group: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
    for (mg, pn), recs in final_group.items():
        # Найдем минимальную цену среди записей с известной страной (ctr != "")
        known = [r for r in recs if r["country"] != ""]
        min_known_price = min((r["price"] for r in known), default=None)
        for r in recs:
            # Если запись без страны и имеется запись со страной, то оставляем её только если её цена <= min_known_price
            if r["country"] == "" and min_known_price is not None and r["price"] > min_known_price:
                continue
            # Используем ключ (mg, pn, r["country"]) для финальной группы
            key = (mg, pn, r["country"])
            adjusted_group.setdefault(key, []).append(r)
    grouped_products = adjusted_group

    # Шаг 4: Применяем сценарии 1-5 для каждой группы (mg, pn, ctr)
    final_product_data: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = {}
    for key, products in grouped_products.items():
        model_group, product_name, country = key
        if not products:
            continue

        # Разделение на предложения с комментариями и без
        with_comments = [p for p in products if p["comment"].strip()]
        without_comments = [p for p in products if not p["comment"].strip()]

        # Находим минимальную цену среди всех предложений
        min_price = min(p["price"] for p in products)

        # Находим все предложения с минимальной ценой
        min_price_offers = [p for p in products if p["price"] == min_price]

        # Разделение минимальных предложений на с комментариями и без
        min_price_with_comments = [p for p in min_price_offers if p["comment"].strip()]
        min_price_without_comments = [p for p in min_price_offers if not p["comment"].strip()]

        if with_comments and without_comments:
            # Сценарии 2 и 4
            min_price_without = min(p["price"] for p in without_comments)
            min_price_with = min(p["price"] for p in with_comments)
            if min_price_with <= min_price_without:
                best_without = [p for p in without_comments if p["price"] == min_price_without]
                best_with = [p for p in with_comments if p["price"] <= min_price_without]
                final_product_data[key] = best_without + best_with
            else:
                best_without = [p for p in without_comments if p["price"] == min_price_without]
                final_product_data[key] = best_without
        elif with_comments and not without_comments:
            final_product_data[key] = with_comments
        elif without_comments and not with_comments:
            min_price_without = min(p["price"] for p in without_comments)
            best_without = [p for p in without_comments if p["price"] == min_price_without]
            final_product_data[key] = best_without
        else:
            final_product_data[key] = products

    # Шаг 5: Формирование итогового списка продуктов
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(dict)
    for (mg, product_name, country), products in final_product_data.items():
        if product_name not in final_product_list[mg]:
            final_product_list[mg][product_name] = []
        for product in products:
            final_product_list[mg][product_name].append({
                "price": product["price"],
                "country": country,
                "suppliers": product["supplier"],
                "comment": product["comment"]
            })

    return final_product_list

def build_final_product_list_for_brand(
    user_data: dict,
    selected_brand: str
) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Обёртка над build_final_product_list:
    - собирает все товары,
    - если selected_brand != "ALL", оставляет только те model_group,
      что принадлежат нужному brand (по PRODUCT_LIBRARY).
    """
    all_products = build_final_product_list(user_data)
    if selected_brand == "ALL":
        return all_products

    # Фильтруем на уровне model_group, исходя из PRODUCT_LIBRARY[selected_brand]
    # Выбираем только те model_group, что есть у данного бренда.
    filtered_products: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    if selected_brand not in PRODUCT_LIBRARY:
        # Если почему-то бренд не существует, вернём пустой словарь
        return {}

    brand_data = PRODUCT_LIBRARY[selected_brand]  # Например, {"iPhone 14": {}, "iPhone 13": {}}
    for model_group in brand_data.keys():
        if model_group in all_products:
            filtered_products[model_group] = all_products[model_group]

    return filtered_products

async def send_csv(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    file_prefix: str = "best_"  # по умолчанию для best
):
    logging.info("Формируем CSV...")

    current_date = datetime.now().strftime("%d_%m_%Y")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8") as temp_file:
        csv_filename = temp_file.name
        csv_writer = csv.writer(temp_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        main_header = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий"]
        csv_writer.writerow(main_header)

        for brand in PRODUCT_LIBRARY:
            for model_group in PRODUCT_LIBRARY[brand]:
                if model_group not in final_product_list:
                    continue

                csv_writer.writerow([model_group, "Цена", "Страны", "Поставщик", "Комментарий"])

                products = final_product_list[model_group]
                flattened_products = [
                    (product_name, entry)
                    for product_name, entries in products.items()
                    for entry in entries
                ]

                sorted_products = sorted(
                    flattened_products,
                    key=lambda x: (
                        get_sort_key(brand, model_group, x[0]),
                        x[1]["price"]
                    )
                )

                for product_name, entry in sorted_products:
                    csv_writer.writerow([
                        product_name,
                        entry["price"],
                        entry["country"],
                        entry["suppliers"],
                        entry["comment"]
                    ])
                csv_writer.writerow([])

    file_name = f"{file_prefix}{current_date}.csv"  # используем переданный префикс

    try:
        with open(csv_filename, "rb") as f:
            await message.reply_document(
                document=f,
                filename=file_name,
                caption="Ваш список товаров (CSV, Группированный формат) готов!"
            )
    except Exception as e:
        logging.error(f"Ошибка при отправке CSV: {e}")
        await message.reply_text("Произошла ошибка при отправке CSV.")
    finally:
        if os.path.exists(csv_filename):
            try:
                os.remove(csv_filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления временного файла {csv_filename}: {e}")

async def send_csv_flat(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    file_prefix: str = "best_"
):
    logging.info("Формируем CSV (Плоский формат)...")

    current_date = datetime.now().strftime("%d_%m_%Y")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8") as temp_file:
        csv_filename = temp_file.name
        csv_writer = csv.writer(temp_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        header = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"]
        csv_writer.writerow(header)

        # Итерируемся по PRODUCT_LIBRARY, чтобы восстановить связь бренд/модель
        for brand in PRODUCT_LIBRARY:
            for model_group in PRODUCT_LIBRARY[brand]:
                if model_group not in final_product_list:
                    continue
                products = final_product_list[model_group]
                flattened_products = [
                    (product_name, entry)
                    for product_name, entries in products.items()
                    for entry in entries
                ]
                # Сортировка по get_sort_key и цене
                sorted_products = sorted(
                    flattened_products,
                    key=lambda x: (get_sort_key(brand, model_group, x[0]), x[1]["price"])
                )
                for product_name, entry in sorted_products:
                    csv_writer.writerow([
                        product_name,
                        entry["price"],
                        entry["country"],
                        entry["suppliers"],
                        entry["comment"],
                        model_group,  # Категория – модель
                        brand       # Группа – бренд
                    ])
    file_name = f"{file_prefix}{current_date}.csv"
    try:
        with open(csv_filename, "rb") as f:
            await message.reply_document(
                document=f,
                filename=file_name,
                caption="Ваш список товаров (CSV, Плоский формат) готов!"
            )
    except Exception as e:
        logging.error(f"Ошибка при отправке CSV: {e}")
        await message.reply_text("Произошла ошибка при отправке CSV.")
    finally:
        if os.path.exists(csv_filename):
            try:
                os.remove(csv_filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления временного файла {csv_filename}: {e}")

async def send_excel(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    file_prefix: str = "best_"  # по умолчанию для best
):
    logging.info("Формируем Excel...")

    current_date = datetime.now().strftime("%d_%m_%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Prices"

    headers = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий"]
    for col_idx, val in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=val)

    row_idx = 2
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for brand in PRODUCT_LIBRARY:
        for model_group in PRODUCT_LIBRARY[brand]:
            if model_group not in final_product_list:
                continue

            ws.cell(row=row_idx, column=1, value=model_group)
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = yellow_fill
            row_idx += 1

            flattened_products = [
                (product_name, entry)
                for product_name, entries in final_product_list[model_group].items()
                for entry in entries
            ]

            sorted_products = sorted(
                flattened_products,
                key=lambda x: (
                    get_sort_key(brand, model_group, x[0]),
                    x[1]["price"]
                )
            )

            for product_name, entry in sorted_products:
                ws.cell(row=row_idx, column=1, value=product_name)
                ws.cell(row=row_idx, column=2, value=entry["price"])
                ws.cell(row=row_idx, column=3, value=entry["country"])
                ws.cell(row=row_idx, column=4, value=entry["suppliers"])
                ws.cell(row=row_idx, column=5, value=entry["comment"])
                row_idx += 1

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_filename = tmp.name
    wb.save(excel_filename)

    file_name = f"{file_prefix}{current_date}.xlsx"  # используем переданный префикс


    try:
        with open(excel_filename, "rb") as f:
            await message.reply_document(
                document=f,
                filename=file_name,
                caption="Ваш список товаров (Excel, Группированный формат) готов!"
            )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel: {e}")
        await message.reply_text("Произошла ошибка при отправке Excel.")
    finally:
        if os.path.exists(excel_filename):
            try:
                os.remove(excel_filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления временного файла {excel_filename}: {e}")

async def send_excel_flat(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    file_prefix: str = "best_"  # или "markup_" / "gradmark_" в зависимости от команды
):
    logging.info("Формируем Excel (Плоский формат)...")

    current_date = datetime.now().strftime("%d_%m_%Y")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Prices Flat"

    # Заголовок с дополнительными колонками: Категория и Группа
    headers = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    row_idx = 2

    # Итерируемся по PRODUCT_LIBRARY для сохранения порядка брендов и групп
    for brand in PRODUCT_LIBRARY:
        for model_group in PRODUCT_LIBRARY[brand]:
            if model_group not in final_product_list:
                continue

            products = final_product_list[model_group]
            # "Разворачиваем" словарь в список пар: (product_name, entry)
            flattened_products = [
                (product_name, entry)
                for product_name, entries in products.items()
                for entry in entries
            ]
            # Многокритериальная сортировка: по get_sort_key и цене
            sorted_products = sorted(
                flattened_products,
                key=lambda x: (get_sort_key(brand, model_group, x[0]), x[1]["price"])
            )
            for product_name, entry in sorted_products:
                ws.cell(row=row_idx, column=1, value=product_name)
                ws.cell(row=row_idx, column=2, value=entry["price"])
                ws.cell(row=row_idx, column=3, value=entry["country"])
                ws.cell(row=row_idx, column=4, value=entry["suppliers"])
                ws.cell(row=row_idx, column=5, value=entry["comment"])
                ws.cell(row=row_idx, column=6, value=model_group)  # Категория
                ws.cell(row=row_idx, column=7, value=brand)          # Группа
                row_idx += 1

    # Сохраняем Excel во временный файл
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_filename = tmp.name
    wb.save(excel_filename)

    file_name = f"{file_prefix}{current_date}.xlsx"

    try:
        with open(excel_filename, "rb") as f:
            await message.reply_document(
                document=f,
                filename=file_name,
                caption="Ваш список товаров (Excel, Плоский формат) готов!"
            )
    except Exception as e:
        logging.error(f"Ошибка при отправке Excel: {e}")
        await message.reply_text("Произошла ошибка при отправке Excel.")
    finally:
        if os.path.exists(excel_filename):
            try:
                os.remove(excel_filename)
            except Exception as e:
                logging.warning(f"Ошибка удаления временного файла {excel_filename}: {e}")

def generate_text_messages(
        final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
        PRODUCT_LIBRARY: Dict[str, Dict[str, Any]],
        include_comments: bool
) -> List[str] :
    """
    Генерирует список сообщений (каждое ≤ 4096 символов) в HTML-разметке,
    чтобы их можно было передавать в Telegram с parse_mode="HTML".

    1. Бренд выводится в верхнем регистре, жирным курсивом:
       <b><i>{BRAND}</i></b>

    2. Избегаем ситуации, когда бренд оказывается последней строкой сообщения
       без групп (model_group) после него. Если видим, что не влезает
       "бренд + хотя бы первая группа" в текущий chunk, откатываем
       (не пишем бренд в конец) и начинаем новое сообщение (chunk) с этого бренда и групп.

    3. Если отдельная группа (model_group) слишком велика и не помещается
       даже в пустой chunk, то идёт дробление: часть товаров в одном сообщении,
       часть — в другом. При этом каждый новый кусок начинается с заголовка
       model_group, чтобы пользователь не терял контекст.

    :param final_product_list: итоговый словарь, где ключ – имя модели (model_group),
                               а значение – список товаров (словарей)
    :param PRODUCT_LIBRARY: исходный словарь всех доступных товаров
    :param include_comments: отображать ли товары, имеющие комментарий
    :return: список сообщений (каждое ≤ 4096 символов) с HTML-разметкой
    """

    # --------------------------
    # ШАГ 1. Формируем структуру: [ {brand, model_groups: [ {model_group, block_lines: [..строки..]}, ... ]}, ... ]
    # --------------------------
    brand_blocks = []

    for brand in PRODUCT_LIBRARY :
        model_groups_data = []
        any_model_group_exists = False

        for model_group in PRODUCT_LIBRARY[brand] :
            if model_group not in final_product_list :
                continue

            products_dict = final_product_list[model_group]
            # Проверка: есть ли вообще товары для данной model_group
            if not products_dict :
                continue

            any_model_group_exists = True

            # 1. "Разворачиваем" словарь в список пар (product_name, entry)
            flattened_products = [
                (product_name, entry)
                for product_name, entries in products_dict.items()
                for entry in entries
            ]

            # 2. Многокритериальная сортировка:
            #    - Сначала по результату get_sort_key(brand, model_group, product_name)
            #    - Затем по цене (entry["price"]) — по возрастанию, если названия одинаковые
            sorted_flattened = sorted(
                flattened_products,
                key=lambda x : (
                    get_sort_key(brand, model_group, x[0]),
                    x[1]["price"]
                )
            )

            # 3. Формируем список текстовых строк для вывода
            product_lines = []
            for product_name, entry in sorted_flattened :
                price = entry["price"]
                comment = entry["comment"].strip()
                country_name = entry["country"].strip()

                # Если пользователь не хочет видеть товары с комментариями, а товар имеет комментарий — пропускаем
                if not include_comments and comment :
                    continue

                # Экранируем текстовые поля для HTML
                product_name_html = escape_html(product_name)
                comment_html = escape_html(comment) if comment else ""
                country_emoji = TEXT_TO_FLAG.get(country_name, escape_html(country_name))
                formatted_price = f"{price:,}".replace(",", ".")

                line_str = f"{product_name_html} – {formatted_price} {country_emoji}"
                if include_comments and comment_html :
                    line_str += f" {comment_html}"

                product_lines.append(line_str)

            # Если после фильтрации нечего выводить, пропускаем
            if not product_lines :
                continue

            # 4. Добавляем заголовок model_group и формируем блок
            mg_header = f"<b>{escape_html(model_group)}</b>"
            block_lines = [mg_header, ""] + product_lines + [""]

            model_groups_data.append({
                "model_group" : model_group,
                "block_lines" : block_lines
            })

        if any_model_group_exists and model_groups_data :
            brand_blocks.append({
                "brand" : brand,
                "model_groups" : model_groups_data
            })

    # Если после всей фильтрации вообще нечего выводить:
    if not brand_blocks :
        return []

    # --------------------------
    # ШАГ 2. Подготовим функции для добавления блоков в сообщения
    # --------------------------
    messages: List[str] = []
    current_chunk = ""

    def get_length_if_added(base: str, new_line: str) -> int :
        """Вычисляет длину текста, если к 'base' добавить строку 'new_line' (через перевод строки)."""
        if not base :
            return len(new_line)
        else :
            return len(base) + 1 + len(new_line)

    def finalize_chunk() :
        """Сохраняет current_chunk в messages, обрезая пробелы по краям и делая reset."""
        nonlocal current_chunk
        trimmed = current_chunk.strip()
        if trimmed :
            messages.append(trimmed)
        current_chunk = ""

    def try_add_line(line: str, chunk: str) -> bool :
        """
        Проверяет, влезет ли строка 'line' в текст 'chunk' (через \n),
        не превышая MAX_TELEGRAM_TEXT.
        Возвращает True/False.
        """
        needed_len = get_length_if_added(chunk, line)
        return (needed_len <= MAX_TELEGRAM_TEXT)

    def add_line(line: str) :
        """Добавляет строку line в current_chunk (с переводом строки), гарантируя, что она влезет."""
        nonlocal current_chunk
        if not current_chunk :
            current_chunk = line
        else :
            current_chunk += "\n" + line

    def fits_block_in_chunk(block_lines: List[str], chunk: str) -> bool :
        """
        Проверяет, поместится ли целиком block_lines в текущий chunk (не превышая лимита).
        """
        temp_len = len(chunk)
        sep_count = 0 if not chunk else len(block_lines)  # кол-во переводов строк, если chunk не пуст

        total_needed = temp_len + sum(len(line) for line in block_lines) + sep_count
        return total_needed <= MAX_TELEGRAM_TEXT

    def fits_block_in_empty_chunk(block_lines: List[str]) -> bool :
        """
        Проверяет, поместится ли целиком block_lines в ПУСТОЙ chunk (не превышая лимита).
        """
        if not block_lines :
            return True
        lines_len = sum(len(line) for line in block_lines)
        # При склеивании N строк будет N-1 переводов строки
        lines_len += (len(block_lines) - 1) if len(block_lines) > 1 else 0
        return lines_len <= MAX_TELEGRAM_TEXT

    def add_block_with_split(block_lines: List[str]) :
        """
        Добавляет block_lines в current_chunk, если целиком не влезает — начинает новый message.
        Если даже в пустое сообщение не влезает, дробим block_lines частями (внутри группы).
        При дроблении в каждом новом сообщении **сначала** печатается заголовок группы
        (и, возможно, пустая строка, если она была после заголовка).
        """
        nonlocal current_chunk

        # Пытаемся сразу добавить в текущий chunk (целиком)
        if fits_block_in_chunk(block_lines, current_chunk) :
            # Влезает — добавляем
            for line in block_lines :
                add_line(line)
        else :
            # Не влезает в текущий chunk
            # Закрываем его и пробуем в пустой chunk
            finalize_chunk()

            if fits_block_in_empty_chunk(block_lines) :
                # В пустой chunk всё влезает
                for line in block_lines :
                    add_line(line)
            else :
                # Даже в пустой не влезает => дробим внутри самой группы
                add_block_in_parts(block_lines)

    def add_block_in_parts(block_lines: List[str]) :
        """
        Дробим block_lines на несколько сообщений, каждый раз повторяя «шапку» (первые строки),
        если нужно. Предположим, что первая строка block_lines — это заголовок модели,
        вторая — может быть пустая (для красивого отступа), а всё остальное — товары.
        Чтобы пользователь не терял контекст, мы повторяем заголовок (и пустую строку) в каждом сообщении.
        """
        nonlocal current_chunk
        if not block_lines :
            return

        # Выделяем «шапку» (заголовок, возможно пустую строку)
        #  - первая строка точно заголовок
        #  - вторая строка, если пустая, тоже считаем частью «шапки»
        header_lines = [block_lines[0]]
        idx = 1
        if len(block_lines) > 1 and not block_lines[1].strip() :
            header_lines.append(block_lines[1])
            idx = 2

        body_lines = block_lines[idx :]

        # Начинаем новый chunk
        finalize_chunk()
        current_chunk = ""

        # Сразу добавляем шапку
        for hl in header_lines :
            # Если даже шапка не влезает, придётся пропускать
            if not try_add_line(hl, current_chunk) :
                finalize_chunk()
                # Если совсем не лезет, придётся пропустить (или как-то ещё обработать)
                continue
            add_line(hl)  # реально добавляем

        # Далее «поштучно» добавляем строки body_lines
        for bl in body_lines :
            if not try_add_line(bl, current_chunk) :
                # Не влезает — открываем новый chunk, повторяем шапку
                finalize_chunk()
                current_chunk = ""
                for hl in header_lines :
                    if not try_add_line(hl, current_chunk) :
                        # Ничего не поделать
                        continue
                    add_line(hl)

                # Теперь добавляем сам товар
                if try_add_line(bl, current_chunk) :
                    add_line(bl)
                else :
                    # Даже отдельная строка не влезает — пропустим
                    finalize_chunk()
                    continue
            else :
                # Влезает
                add_line(bl)

    # --------------------------
    # ШАГ 3. Основная логика: добавляем блоки для брендов и групп, избегая "бренда в одиночестве в конце"
    # --------------------------

    last_brand = None

    for brand_block in brand_blocks :
        brand = brand_block["brand"]  # Например, "Apple"
        mg_list = brand_block["model_groups"]  # Список { model_group, block_lines }

        brand_upper = brand.upper()
        brand_header = f"<b><i>{escape_html(brand_upper)}</i></b>"

        # Проходим по каждой model_group для данного бренда.
        # Но сначала надо *убедиться*, что мы не «повесим» бренд в конец текущего chunk,
        # если туда не влезет хотя бы одна группа.

        for idx_mg, mg_item in enumerate(mg_list) :
            mg_block_lines = mg_item["block_lines"]

            # Проверяем, уже ли мы «вставляли» этот brand в текущий chunk
            brand_already_printed = (brand == last_brand and current_chunk.strip() != "")

            # Если мы ещё НЕ печатали этот бренд в текущем chunk, то сначала добавим brand + сразу текущую группу
            if not brand_already_printed :
                # 1. Попробуем «виртуально» добавить brand + mg_block_lines в текущий chunk
                #    чтобы увидеть, влезает ли всё целиком.
                test_lines = [brand_header, ""] + mg_block_lines  # бренд, пустая строка, блок группы

                if fits_block_in_chunk(test_lines, current_chunk) :
                    # Всё влезает
                    # Добавим реально
                    for line in test_lines :
                        add_line(line)
                    last_brand = brand
                else :
                    # Не влезает в текущий chunk
                    # => «откатываем» бренд (не добавляли же фактически) и начинаем новый chunk
                    finalize_chunk()

                    # Пробуем в пустом chunk
                    if fits_block_in_empty_chunk(test_lines) :
                        # Поместится
                        for line in test_lines :
                            add_line(line)
                        last_brand = brand
                    else :
                        # Даже в пустой chunk целиком не помещается => нужно дробить
                        # Но дробить мы будем всё вместе: brand_header + "" + mg_block_lines
                        # Если brand_header + "" + mg_block_lines тоже не помещается,
                        # внутри придётся делить. Однако бренд — тоже часть «шапки» для первой группы.
                        # Для упрощения пойдём через ту же логику add_block_in_parts:
                        big_block = [brand_header, ""] + mg_block_lines
                        add_block_in_parts(big_block)
                        last_brand = brand
            else :
                # Бренд уже напечатан в текущем chunk, значит просто добавляем очередную группу
                # (целиком или с дроблением, если не влезает)
                add_block_with_split(mg_block_lines)

        # По окончании бренда можно (по желанию) добавить пустую строку или нет.
        # Но главное — если следующий бренд пойдёт, он уже проверит свои условия.

    # В самом конце — финализируем текущий chunk, если что-то накопилось
    finalize_chunk()

    return messages

# ------------------------------ ГЛАВНЫЕ ОБРАБОТЧИКИ BEST------------------------------
# Новые callback-паттерны для выбора структуры таблицы
STRUCTURE_GROUPED = "structure_grouped"
STRUCTURE_FLAT = "structure_flat"

# Добавляем новое состояние (например, для /best можно определить его как)
ASK_STRUCTURE = 6   # новое состояние для выбора структуры
ASK_COMMENTS  = 7   # сдвигаем состояние вопросов про комментарии

def build_final_product_list_for_multiple_brands(
        user_data: dict,
        brand_list: List[str]
) -> Dict[str, Dict[str, List[Dict[str, Any]]]] :
    """
    Если пользователь выбрал несколько брендов разом (например, «Apple iPhone» и «Apple Watch»),
    то нужно объединить результаты из build_final_product_list_for_brand
    для каждого бренда в brand_list.

    Возвращает общий словарь final_product_list,
    скомбинированный из нескольких «под-брендов».
    """

    if not brand_list :
        return {}

    combined_result: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}

    for single_brand in brand_list :
        # вызываем вашу существующую функцию
        sub_result = build_final_product_list_for_brand(user_data, single_brand)

        # теперь сливаем sub_result в общий combined_result
        for model_group, product_dict in sub_result.items() :
            if model_group not in combined_result :
                combined_result[model_group] = {}
            for product_name, entries in product_dict.items() :
                if product_name not in combined_result[model_group] :
                    combined_result[model_group][product_name] = []
                # Добавляем все товары
                combined_result[model_group][product_name].extend(entries)

    return combined_result

def build_final_product_list_for_all(user_data: dict) -> Dict[str, Dict[str, List[Dict[str, Any]]]] :
    """
    Если пользователь выбрал «Все группы», берем вообще все товары
    (аналог brand="ALL").
    """
    return build_final_product_list(user_data)

async def best_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # 3) /best: Шаг 1 — Выбор «Все группы» или одна из LIST_BRAND_GROUP
    """
    Шаг 1: показать "Все группы" + группы из LIST_BRAND_GROUP.
    Пример:
     - "Все группы"
     - "Apple"
     - "Устройства на Android"
     - "Dyson"
    """
    user_id = update.effective_user.id

    # Проверяем, есть ли у пользователя товары
    user_products = get_all_products(user_id)
    if not user_products :
        await update.message.reply_text("Ваш список товаров пуст. Добавьте товары...")
        return

    # Формируем кнопку "Все группы"
    keyboard = []
    keyboard.append([InlineKeyboardButton("Все группы", callback_data="group_ALL")])

    # Кнопки для каждой ключа в LIST_BRAND_GROUP
    for group_name in LIST_BRAND_GROUP.keys() :
        callback_data = f"group_{group_name}"
        keyboard.append([InlineKeyboardButton(group_name, callback_data=callback_data)])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите группу:", reply_markup=reply_markup)

async def best_group_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) :
    # 4) best_group_callback: обрабатывает нажатие «group_ALL» или «group_Apple»
    """
    Если group_ALL => переходим сразу к выбору формата (CSV/Excel/Сообщение).
    Иначе — показываем «Все категории» + список sub-брендов.
    """
    query = update.callback_query
    data = query.data  # "group_ALL" или "group_Apple", "group_Устройства на Android", "group_Dyson", ...
    await query.answer()

    if data == "group_ALL" :
        # Сохраняем в user_data, что выбрана "ALL" (все бренды)
        context.user_data["selected_group"] = "ALL"
        # Раз сразу "Все группы", идём прямо к выбору формата
        keyboard = [
            [
                InlineKeyboardButton("CSV", callback_data="best_csv"),
                InlineKeyboardButton("Excel", callback_data="best_excel"),
                InlineKeyboardButton("Сообщение", callback_data="best_msg"),
            ]
        ]
        await query.edit_message_text("Выберите формат:", reply_markup=InlineKeyboardMarkup(keyboard))
        return

    # Иначе выбрана конкретная группа (например, "Apple")
    group_name = data.replace("group_", "")
    context.user_data["selected_group"] = group_name

    # Формируем меню «Все категории» + список sub-брендов
    sub_brands = LIST_BRAND_GROUP.get(group_name, [])

    keyboard = []
    # "Все категории"
    keyboard.append([InlineKeyboardButton("Все категории", callback_data="cat_ALL")])
    # каждая sub-бренд
    for sb in sub_brands :
        cb = f"cat_{sb}"
        keyboard.append([InlineKeyboardButton(sb, callback_data=cb)])

    await query.edit_message_text("Выберите категорию:", reply_markup=InlineKeyboardMarkup(keyboard))

async def best_category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) :
    # 5) best_category_callback: обрабатывает «cat_ALL» или «cat_Apple iPhone»
    """
    Если cat_ALL => выбраны все под-бренды группы.
    Если cat_XXX => выбран конкретный под-бренд.
    Переходим к шагу "выберите формат".
    """
    query = update.callback_query
    data = query.data  # "cat_ALL" или "cat_Apple iPhone", ...
    await query.answer()

    group_name = context.user_data.get("selected_group", "ALL")

    if data == "cat_ALL" :
        # Тогда пользователь берет все sub-бренды этой группы
        sub_brands = LIST_BRAND_GROUP.get(group_name, [])
        context.user_data["selected_brands"] = sub_brands
    else :
        # Выбрана конкретная категория, например "cat_Apple iPhone"
        brand_str = data.replace("cat_", "")
        context.user_data["selected_brands"] = [brand_str]

    # Теперь показываем формат (CSV / Excel / Сообщение)
    keyboard = [
        [
            InlineKeyboardButton("CSV", callback_data="best_csv"),
            InlineKeyboardButton("Excel", callback_data="best_excel"),
            InlineKeyboardButton("Сообщение", callback_data="best_msg"),
        ]
    ]
    await query.edit_message_text("Выберите формат:", reply_markup=InlineKeyboardMarkup(keyboard))

async def best_command_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) :
    query = update.callback_query
    choice = query.data  # "best_csv", "best_excel" или "best_msg"
    await query.answer()

    # Загружаем данные из SQL и формируем итоговый список
    user_id = update.effective_user.id
    pseudo_data = load_pseudo_user_data(user_id)
    group_name = context.user_data.get("selected_group", "ALL")
    selected_sub_brands = context.user_data.get("selected_brands", [])
    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_sub_brands)

    if not final_product_list :
        await query.edit_message_text("Данные не найдены (пусто?).")
        return

    context.user_data["best_final_product_list"] = final_product_list
    context.user_data["best_chosen_format"] = choice

    # Если выбран формат CSV или Excel, спрашиваем структуру таблицы
    if choice in ["best_csv", "best_excel"] :
        keyboard = [
            [
                InlineKeyboardButton("Группированный формат", callback_data=STRUCTURE_GROUPED),
                InlineKeyboardButton("Плоский формат", callback_data=STRUCTURE_FLAT),
            ]
        ]
        await query.edit_message_text(
            text="Выберите структуру данных таблицы:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_STRUCTURE  # переходим в новое состояние выбора структуры

    # Если выбран формат "Сообщение" – переходим к вопросу "Учитывать комментарии?"
    else :
        keyboard = [
            [
                InlineKeyboardButton("Да", callback_data="best_msg_comments_yes"),
                InlineKeyboardButton("Нет", callback_data="best_msg_comments_no"),
            ]
        ]
        await query.edit_message_text(
            text="Нужно ли учитывать комментарии?",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_COMMENTS

async def best_structure_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    structure_choice = query.data  # "structure_grouped" или "structure_flat"
    await query.answer()

    # Сохраняем выбор структуры
    if structure_choice == STRUCTURE_GROUPED:
        context.user_data["best_structure"] = "grouped"
    else:
        context.user_data["best_structure"] = "flat"

    # После выбора структуры спрашиваем: "Нужно ли учитывать комментарии?"
    keyboard = [
        [
            InlineKeyboardButton("Да", callback_data="best_msg_comments_yes"),
            InlineKeyboardButton("Нет", callback_data="best_msg_comments_no"),
        ]
    ]
    await query.edit_message_text(
        text="Нужно ли учитывать комментарии?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ASK_COMMENTS

async def best_command_comments_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    choice = query.data  # "best_msg_comments_yes" или "best_msg_comments_no"
    await query.answer()

    include_comments = (choice == "best_msg_comments_yes")

    chosen_format = context.user_data.get("best_chosen_format", "best_msg")
    final_product_list = context.user_data.get("best_final_product_list", {})

    if not final_product_list:
        await query.edit_message_text("Не удалось сформировать данные (пусто?).")
        return

    # Удаляем кнопки в любом случае
    await query.edit_message_reply_markup(reply_markup=None)

    if chosen_format == "best_csv":
        if not include_comments:
            remove_commented_products(final_product_list)
        await query.message.reply_text("📑 Генерируем CSV...")
        if context.user_data.get("best_structure", "grouped") == "flat":
            await send_csv_flat(query.message, context, final_product_list, file_prefix="best_")
        else:
            await send_csv(query.message, context, final_product_list, file_prefix="best_")
    elif chosen_format == "best_excel":
        if not include_comments:
            remove_commented_products(final_product_list)
        await query.message.reply_text("📑 Генерируем Excel...")
        if context.user_data.get("best_structure", "grouped") == "flat":
            await send_excel_flat(query.message, context, final_product_list, file_prefix="best_")
        else:
            await send_excel(query.message, context, final_product_list, file_prefix="best_")
    elif chosen_format == "best_msg":
        messages = generate_text_messages(final_product_list, PRODUCT_LIBRARY, include_comments=include_comments)
        for msg in messages:
            await query.message.reply_text(msg, parse_mode="HTML")
    else:
        await query.message.reply_text("Неизвестный формат.")

# ------------------------------ ПЕРЕМЕННЫЕ ФУНКЦИИ ДЛЯ ФОРМИРОВАНИЯ markup И ТОЧКА ВХОДА------------------------------

# STATES (ConversationHandler)
# Новые константы для выбора структуры в рамках /markup
MARKUP_STRUCTURE_GROUPED = "markup_structure_grouped"
MARKUP_STRUCTURE_FLAT = "markup_structure_flat"

markup_TYPE = 1      # Выбор "фикс/процент"
ENTER_VALUE = 2      # Ввод числа
CHOOSE_GROUP = 3     # Выбор группы (Apple / Dyson / "Все группы")
CHOOSE_CATEGORY = 4  # Выбор категории (внутри группы)
CHOOSE_FORMAT = 5    # Выбор CSV/Excel/Сообщение
ASK_COMMENT = 6     # Да/Нет по комментариям
# Добавляем новое состояние для выбора структуры
ASK_MARKUP_STRUCTURE = 7
# Сдвигаем состояние для вопроса о комментариях:
ASK_MARKUP_COMMENT = 8

# CALLBACK PATTERNS для типов наценки
markup_FIXED = "markup_fixed"
markup_PERCENT = "markup_percent"

FORMAT_CSV = "markup_csv"
FORMAT_EXCEL = "markup_excel"
FORMAT_MSG = "markup_msg"

MSG_COMMENTS_YES = "markup_msg_comments_yes"
MSG_COMMENTS_NO  = "markup_msg_comments_no"

# Лимиты
MAX_markup_VALUE = 1_000_000  # рублей
MAX_markup_PERCENT = 1_000    # 1000%

# Список сообщений об ошибках для этапа ввода наценки
ERROR_MESSAGES_MARKUP_FORMAT = [
    "🔄 Ой! Похоже, формат ввода немного не соответствует примеру. Пожалуйста, проверьте и попробуйте снова. 😊",
    "📝 Кажется, в вашем вводе есть неточности. Убедитесь, что вы следуете примеру, и повторите попытку. 🚀",
    "❗ Внимание! Формат данных не распознан. Пожалуйста, введите градацию цен согласно образцу. 💡",
    "🤔 Хм, что-то не так. Введите данные в правильном формате. 📋",
    "🔢 Похоже, числа или диапазоны указаны неверно. Проверьте формат и попробуйте ещё раз. 🧮",
    "🛠 Упс! Возникла ошибка в формате ввода. Давайте сверимся с примером и повторим попытку. 📝",
    "📊 Данные не соответствуют формату. Используйте пример в качестве ориентира. 😊",
    "🙈 Я не смог обработать ваш ввод. Убедитесь, что вы следуете структуре примера, и попробуйте снова. 🔄",
    "🚧 Ошибка! Формат градации цен некорректен. Давайте проверим ещё раз по примеру и повторим. 💪",
    "🔍 Не удаётся распознать введённые данные. Пожалуйста, внимательно следуйте примеру и попробуйте снова. 📖"
]

# ------------------------------ ВХОДНАЯ ТОЧКА: /markup ------------------------------

async def markup_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Начало сценария /markup.
    Шаг 1: выбор "Твердая наценка" / "Процентная наценка".
    Переходим в состояние markup_TYPE.
    """
    user_id = update.effective_user.id
    user_products = get_all_products(user_id)
    if not user_products :
        await update.message.reply_text("У вас нет товаров — нечего наценивать! Сначала добавьте товары.")
        return ConversationHandler.END

    keyboard = [
        [InlineKeyboardButton("Твердая наценка (руб.)", callback_data=markup_FIXED)],
        [InlineKeyboardButton("Наценка в процентах (%)", callback_data=markup_PERCENT)],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        text="Выберите способ наценки:",
        reply_markup=reply_markup
    )
    return markup_TYPE

# ------------------------------ ГЛАВНЫЕ ОБРАБОТЧИКИ markup------------------------------

async def markup_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Пользователь нажал "Твердая наценка" / "Наценка %".
    Сохраняем в user_data["markup_type"], просим ввести число.
    """
    query = update.callback_query
    data = query.data  # markup_FIXED / markup_PERCENT
    await query.answer()

    if data == markup_FIXED:
        context.user_data["markup_type"] = "fixed"
        await query.edit_message_text("Введите сумму наценки (руб.):")
    elif data == markup_PERCENT:
        context.user_data["markup_type"] = "percent"
        await query.edit_message_text("Введите процент наценки (например, 10 или 10%).\nОкругление будет до ближайших 100 рублей.")
    else:
        await query.edit_message_text("Неизвестный тип наценки.")
        return ConversationHandler.END

    return ENTER_VALUE

async def markup_value_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Пользователь вводит число (или число с %).
    Парсим, проверяем лимиты, если ок => переходим к выбору "группы" (CHOOSE_GROUP).
    Если формат неверный – отправляем одно из случайных сообщений об ошибке.
    """
    user_input = update.message.text.strip().lower()
    markup_type = context.user_data.get("markup_type", "fixed")

    # Ищем число в строке
    just_digits = re.findall(r"\d+(?:[.,]\d+)?", user_input)
    if not just_digits:
        return await _markup_retry_or_abort(update, context)

    raw_number = just_digits[0].replace(",", ".")
    try:
        value = float(raw_number)
    except ValueError:
        return await _markup_retry_or_abort(update, context)

    # Проверяем лимиты
    if markup_type == "fixed":
        if value > MAX_markup_VALUE:
            return await _markup_retry_or_abort(update, context)
        context.user_data["markup_value"] = value
        await update.message.reply_text(
            f"Принята твёрдая наценка: {value:.0f} руб.\nТеперь выберите группу:"
        )
    else:  # "percent"
        if value > MAX_markup_PERCENT:
            return await _markup_retry_or_abort(update, context)
        context.user_data["markup_value"] = value
        await update.message.reply_text(
            f"Принята наценка: {value:.1f}%\nТеперь выберите группу:"
        )

    # Переходим к состоянию CHOOSE_GROUP: отправляем кнопки "Все группы" и список групп (LIST_BRAND_GROUP)
    keyboard = []
    keyboard.append([InlineKeyboardButton("Все группы", callback_data="group_ALL")])
    for group_name in LIST_BRAND_GROUP.keys():
        callback_data = f"group_{group_name}"
        keyboard.append([InlineKeyboardButton(group_name, callback_data=callback_data)])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите группу:", reply_markup=reply_markup)
    return CHOOSE_GROUP

async def _markup_retry_or_abort(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Вспомогательная функция для обработки неверного формата ввода на этапе markup.
    Увеличивает счетчик попыток и либо просит ввести снова, либо завершает сценарий после 3 ошибок.
    """
    attempts = context.user_data.get("markup_attempts", 0) + 1
    context.user_data["markup_attempts"] = attempts
    error_msg = random.choice(ERROR_MESSAGES_MARKUP_FORMAT)
    if attempts >= 3:
        await update.message.reply_text(f"{error_msg}\nСлишком много ошибок. Процесс прерван.")
        return ConversationHandler.END
    else:
        await update.message.reply_text(error_msg)
        return ENTER_VALUE

async def markup_group_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Аналог best_group_callback:
      - Если group_ALL => переходим сразу к выбору формата
      - Если group_Apple => показываем "Все категории" + sub-brand'ы
    """
    query = update.callback_query
    data = query.data  # "group_ALL" / "group_Apple" / ...
    await query.answer()

    if data == "group_ALL":
        # user выбрал "Все группы"
        context.user_data["selected_group"] = "ALL"

        # Сразу к выбору формата
        keyboard = [
            [
                InlineKeyboardButton("CSV", callback_data=FORMAT_CSV),
                InlineKeyboardButton("Excel", callback_data=FORMAT_EXCEL),
                InlineKeyboardButton("Сообщение", callback_data=FORMAT_MSG),
            ]
        ]
        await query.edit_message_text("Выберите формат:", reply_markup=InlineKeyboardMarkup(keyboard))
        return CHOOSE_FORMAT

    # Иначе пользователь выбрал конкретную группу
    group_name = data.replace("group_", "")
    context.user_data["selected_group"] = group_name

    # Показываем "Все категории" + список sub-брендов
    sub_brands = LIST_BRAND_GROUP.get(group_name, [])

    keyboard = []
    keyboard.append([InlineKeyboardButton("Все категории", callback_data="cat_ALL")])
    for sb in sub_brands:
        cb = f"cat_{sb}"
        keyboard.append([InlineKeyboardButton(sb, callback_data=cb)])

    await query.edit_message_text("Выберите категорию:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_CATEGORY

async def markup_category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Аналог best_category_callback:
      - Если cat_ALL => user выбрал все под-бренды
      - Иначе => один под-бренд
    Переход к выбору формата (CSV/Excel/Сообщение).
    """
    query = update.callback_query
    data = query.data
    await query.answer()

    group_name = context.user_data.get("selected_group", "ALL")

    if data == "cat_ALL":
        # все sub-бренды
        sub_brands = LIST_BRAND_GROUP.get(group_name, [])
        context.user_data["selected_brands"] = sub_brands
    else:
        # например "cat_Apple iPhone" => "Apple iPhone"
        brand_str = data.replace("cat_", "")
        context.user_data["selected_brands"] = [brand_str]

    # Переходим к выбору формата
    keyboard = [
        [
            InlineKeyboardButton("CSV", callback_data=FORMAT_CSV),
            InlineKeyboardButton("Excel", callback_data=FORMAT_EXCEL),
            InlineKeyboardButton("Сообщение", callback_data=FORMAT_MSG),
        ]
    ]
    await query.edit_message_text("Выберите формат:", reply_markup=InlineKeyboardMarkup(keyboard))
    return CHOOSE_FORMAT

async def markup_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка нажатия CSV/Excel/Сообщение при наценке.
    Теперь: если выбран CSV или Excel, сначала загружаем данные и спрашиваем о структуре таблицы.
    """
    query = update.callback_query
    choice = query.data  # FORMAT_CSV, FORMAT_EXCEL или FORMAT_MSG
    await query.answer()

    context.user_data["markup_chosen_format"] = choice

    if choice in [FORMAT_CSV, FORMAT_EXCEL]:
        # Загружаем данные и формируем итоговый список товаров
        user_id = update.effective_user.id
        pseudo_data = load_pseudo_user_data(user_id)
        group_name = context.user_data.get("selected_group", "ALL")
        selected_sub_brands = context.user_data.get("selected_brands", [])
        if group_name == "ALL":
            final_product_list = build_final_product_list_for_all(pseudo_data)
        else:
            final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_sub_brands)
        context.user_data["markup_final_product_list"] = final_product_list

        # Спросим, в каком формате выводить таблицу
        keyboard = [
            [
                InlineKeyboardButton("Группированный формат", callback_data=MARKUP_STRUCTURE_GROUPED),
                InlineKeyboardButton("Плоский формат", callback_data=MARKUP_STRUCTURE_FLAT),
            ]
        ]
        await query.edit_message_text(
            text="Выберите структуру данных таблицы:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_MARKUP_STRUCTURE
    else:
        # Для формата "Сообщение" сразу спрашиваем про комментарии
        keyboard = [
            [
                InlineKeyboardButton("Да", callback_data=MSG_COMMENTS_YES),
                InlineKeyboardButton("Нет", callback_data=MSG_COMMENTS_NO),
            ]
        ]
        await query.edit_message_text(
            text="Учитывать комментарии?",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_MARKUP_COMMENT

async def markup_structure_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка выбора структуры таблицы для /markup.
    Сохраняем выбор и переходим к вопросу об учёте комментариев.
    """
    query = update.callback_query
    structure_choice = query.data  # MARKUP_STRUCTURE_GROUPED или MARKUP_STRUCTURE_FLAT
    await query.answer()

    if structure_choice == MARKUP_STRUCTURE_GROUPED:
        context.user_data["markup_structure"] = "grouped"
    else:
        context.user_data["markup_structure"] = "flat"

    keyboard = [
        [
            InlineKeyboardButton("Да", callback_data=MSG_COMMENTS_YES),
            InlineKeyboardButton("Нет", callback_data=MSG_COMMENTS_NO),
        ]
    ]
    await query.edit_message_text(
        text="Учитывать комментарии?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ASK_MARKUP_COMMENT

async def markup_comments_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка ответа на вопрос об учёте комментариев для /markup.
    В зависимости от выбранного формата и структуры вызываем соответствующую функцию.
    """
    query = update.callback_query
    choice = query.data  # MSG_COMMENTS_YES или MSG_COMMENTS_NO
    await query.answer()

    include_comments = (choice == MSG_COMMENTS_YES)
    chosen_format = context.user_data.get("markup_chosen_format")

    if chosen_format == FORMAT_CSV:
        if context.user_data.get("markup_structure", "grouped") == "flat":
            await query.edit_message_text("Генерируем CSV с наценкой (Плоский формат)...")
            await send_csv_flat_with_markup(query, context, include_comments=include_comments)
        else:
            await query.edit_message_text("Генерируем CSV с наценкой...")
            await send_csv_with_markup(query, context, include_comments=include_comments)
        return ConversationHandler.END
    elif chosen_format == FORMAT_EXCEL:
        if context.user_data.get("markup_structure", "grouped") == "flat":
            await query.edit_message_text("Генерируем Excel с наценкой (Плоский формат)...")
            await send_excel_flat_with_markup(query, context, include_comments=include_comments)
        else:
            await query.edit_message_text("Генерируем Excel с наценкой...")
            await send_excel_with_markup(query, context, include_comments=include_comments)
        return ConversationHandler.END
    elif chosen_format == FORMAT_MSG:
        await query.edit_message_text("Формируем сообщение с наценкой...")
        await send_message_with_markup(query, context, include_comments=include_comments)
        return ConversationHandler.END
    else:
        await query.edit_message_text("Неизвестный формат.")
        return ConversationHandler.END

# ------------------------------ ФУНКЦИИ ДЛЯ ФОРМИРОВАНИЯ markup------------------------------

async def send_csv_with_markup(query, context, include_comments=False) :
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    sub_brands = context.user_data.get("selected_brands", [])
    markup_type = context.user_data.get("markup_type", "fixed")
    markup_value = context.user_data.get("markup_value", 0)
    pseudo_data = load_pseudo_user_data(user_id)

    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, sub_brands)

    # Применяем наценку для обновления цен
    apply_markup_to_final_list(final_product_list, markup_type, markup_value)

    if not include_comments :
        remove_commented_products(final_product_list)

    # Передаём префикс "markup_csv_" вместо стандартного
    await send_csv(query.message, context, final_product_list, file_prefix="markup_")

async def send_csv_flat_with_markup(query, context, include_comments=False):
    """
    Генерирует CSV-файл для /markup в плоском формате с наценкой.
    """
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    sub_brands = context.user_data.get("selected_brands", [])
    markup_type = context.user_data.get("markup_type", "fixed")
    markup_value = context.user_data.get("markup_value", 0)

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, sub_brands)

    apply_markup_to_final_list(final_product_list, markup_type, markup_value)
    if not include_comments:
        remove_commented_products(final_product_list)

    await send_csv_flat(query.message, context, final_product_list, file_prefix="markup_")

async def send_excel_with_markup(query, context, include_comments=False):
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    sub_brands = context.user_data.get("selected_brands", [])

    markup_type = context.user_data.get("markup_type", "fixed")
    markup_value = context.user_data.get("markup_value", 0)

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, sub_brands)

    apply_markup_to_final_list(final_product_list, markup_type, markup_value)

    if not include_comments:
        remove_commented_products(final_product_list)

    # Передаём префикс "markup_excel_" вместо стандартного
    await send_excel(query.message, context, final_product_list, file_prefix="markup_")

async def send_excel_flat_with_markup(query, context, include_comments=False):
    """
    Генерирует Excel-файл для /markup в плоском формате с наценкой.
    """
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    sub_brands = context.user_data.get("selected_brands", [])
    markup_type = context.user_data.get("markup_type", "fixed")
    markup_value = context.user_data.get("markup_value", 0)

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, sub_brands)

    apply_markup_to_final_list(final_product_list, markup_type, markup_value)
    if not include_comments:
        remove_commented_products(final_product_list)

    await send_excel_flat(query.message, context, final_product_list, file_prefix="markup_")

async def send_message_with_markup(query, context, include_comments=False):
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    sub_brands = context.user_data.get("selected_brands", [])

    markup_type = context.user_data.get("markup_type", "fixed")
    markup_value = context.user_data.get("markup_value", 0)

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, sub_brands)

    apply_markup_to_final_list(final_product_list, markup_type, markup_value)

    messages = generate_text_messages(
        final_product_list,
        PRODUCT_LIBRARY,
        include_comments=include_comments
    )
    for msg in messages :
        await query.message.reply_text(msg, parse_mode="HTML")

def round_up_to_nearest_100(amount) :
    """
    Округляет сумму до ближайших 100 рублей вверх при выборе наценки в %.

    Примеры:
    1234 -> 1300
    1200 -> 1200
    """
    return math.ceil(amount / 100.0) * 100

def apply_markup_to_final_list(final_product_list, markup_type, markup_value) :
    """
    Проходит по всем товарам и к price прибавляет наценку.
    При процентной наценке округляет цену до ближайших 100 рублей вверх.

    markup_type: "fixed" или "percent"
    markup_value: float
    """
    if not final_product_list :
        return

    for model_group, product_dict in final_product_list.items() :
        for product_name, entries in product_dict.items() :
            for entry in entries :
                orig_price = entry["price"]
                if markup_type == "fixed" :
                    new_price = orig_price + markup_value
                else :
                    # Применяем процентную наценку
                    new_price = orig_price + (orig_price * (markup_value / 100.0))
                    # Округляем вверх до ближайших 100 рублей
                    new_price = round_up_to_nearest_100(new_price)

                # Округляем до целого числа рублей
                entry["price"] = int(new_price)

def get_markup_conversation_handler():
    return ConversationHandler(
        entry_points=[CommandHandler("markup", markup_command)],
        states={
            markup_TYPE: [
                CallbackQueryHandler(markup_type_callback, pattern=f"^{markup_FIXED}|{markup_PERCENT}$")
            ],
            ENTER_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, markup_value_handler)
            ],
            CHOOSE_GROUP: [
                CallbackQueryHandler(markup_group_callback, pattern=r"^group_")
            ],
            CHOOSE_CATEGORY: [
                CallbackQueryHandler(markup_category_callback, pattern=r"^cat_")
            ],
            CHOOSE_FORMAT: [
                CallbackQueryHandler(markup_format_callback, pattern=f"^{FORMAT_CSV}|{FORMAT_EXCEL}|{FORMAT_MSG}$")
            ],
            ASK_MARKUP_STRUCTURE: [
                CallbackQueryHandler(markup_structure_callback, pattern=f"^{MARKUP_STRUCTURE_GROUPED}|{MARKUP_STRUCTURE_FLAT}$")
            ],
            ASK_MARKUP_COMMENT: [
                CallbackQueryHandler(markup_comments_callback, pattern=f"^{MSG_COMMENTS_YES}|{MSG_COMMENTS_NO}$")
            ],
        },
        fallbacks=[
            CommandHandler("cancel", lambda u, c: ConversationHandler.END),
        ]
    )


# ------------------------------ Константы и STATES для gradmark ------------------------------

# Константы для градационной наценки
gradmark_TYPE = 1            # Выбор типа градационной наценки (фиксированная или процентная)
ENTER_GRADATION = 2          # Ввод градационных диапазонов
ASK_GRADMARK_STRUCTURE = 6   # Новый шаг: выбор структуры таблицы

# Callback-паттерны для градационной наценки
gradmark_FIXED = "gradmark_fixed"
gradmark_PERCENT = "gradmark_percent"

# Новые callback-паттерны для выбора структуры в градационной наценке
GRADMARK_STRUCTURE_GROUPED = "gradmark_structure_grouped"
GRADMARK_STRUCTURE_FLAT = "gradmark_structure_flat"

# Список сообщений об ошибках для этапа градации цен
ERROR_MESSAGES_GRADMARK_FORMAT = [
    "🔄 Ой! Похоже, формат ввода немного не соответствует примеру. Пожалуйста, проверьте и попробуйте снова. 😊",
    "📝 Кажется, в вашем вводе есть неточности. Убедитесь, что вы следуете примеру, и повторите попытку. 🚀",
    "❗ Внимание! Формат данных не распознан. Пожалуйста, введите градацию цен согласно образцу. 💡",
    "🤔 Хм, что-то не так. Введите данные в правильном формате. 📋",
    "🔢 Похоже, числа или диапазоны указаны неверно. Проверьте формат и попробуйте ещё раз. 🧮",
    "🛠 Упс! Возникла ошибка в формате ввода. Давайте сверимся с примером и повторим попытку. 📝",
    "📊 Данные не соответствуют формату. Используйте пример в качестве ориентира. 😊",
    "🙈 Я не смог обработать ваш ввод. Убедитесь, что вы следуете структуре примера, и попробуйте снова. 🔄",
    "🚧 Ошибка! Формат градации цен некорректен. Давайте проверим ещё раз по примеру и повторим. 💪",
    "🔍 Не удаётся распознать введённые данные. Пожалуйста, внимательно следуйте примеру и попробуйте снова. 📖"
]

# ------------------------------ Функции для команды /gradmark ------------------------------

async def gradmark_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int :
    """
    Точка входа в сценарий /gradmark.
    Шаг 1: выбор типа градационной наценки (фиксированная или процентная).
    """
    user_id = update.effective_user.id
    # Проверяем, есть ли товары у пользователя
    rows = get_all_products(user_id)
    if not rows :
        await update.message.reply_text("У вас пока нет товаров. Сначала добавьте товары!")
        return ConversationHandler.END

    keyboard = [
        [InlineKeyboardButton("Твердая наценка (руб.)", callback_data=gradmark_FIXED)],
        [InlineKeyboardButton("Наценка в процентах (%)", callback_data=gradmark_PERCENT)],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        text="Выберите способ градационной наценки:",
        reply_markup=reply_markup
    )
    return gradmark_TYPE

async def gradmark_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int :
    """
    Обработка выбора типа градационной наценки.
    Сохраняем тип (fixed или percent) и отправляем сообщение с примером ввода градации.
    """
    query = update.callback_query
    data = query.data
    await query.answer()

    if data == gradmark_FIXED :
        context.user_data["gradmark_type"] = "fixed"
        text = (
            "Отлично! Теперь приступим к настройке градации цен с твердой наценкой:\n\n"
            "📋 Пример ввода для:\n\n"
            "0-10000 1000\n"
            "10001-50000 1500\n"
            "50001-100000 3000\n"
            "100001+ 7500\n\n"
            "Объяснение ввода:\n"
            "Первый диапазон цен начинается с 0 до N.\n"
            "Далее идут последовательные диапазоны цен – каждый диапазон должен логически следовать за предыдущим, без пропусков чисел.\n"
            "Пример: 10001-50000 следует за 10000.\n"
            "Последняя строка с N+ означает, что указанная наценка применяется ко всем суммам, начиная с N и выше, без ограничения верхнего предела."
        )
    elif data == gradmark_PERCENT :
        context.user_data["gradmark_type"] = "percent"
        text = (
            "Отлично! Теперь приступим к настройке градации цен с процентной наценкой:\n\n"
            "📋 Пример ввода для:\n\n"
            "0-10000 1.5%\n"
            "10001-50000 3%\n"
            "50001-100000 5%\n"
            "100001+ 7%\n\n"
            "Объяснение ввода:\n"
            "Первый диапазон цен начинается с 0 до N.\n"
            "Далее идут последовательные диапазоны цен – каждый диапазон должен логически следовать за предыдущим, без пропусков чисел.\n"
            "Пример: 10001-50000 следует за 10000.\n"
            "Последняя строка с N+ означает, что указанная наценка применяется ко всем суммам, начиная с N и выше, без ограничения верхнего предела."
        )
    else :
        await query.edit_message_text("Неизвестный тип градационной наценки.")
        return ConversationHandler.END

    await query.edit_message_text(text)
    return ENTER_GRADATION

async def gradmark_gradation_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int :
    """
    Обрабатывает ввод градационных диапазонов.
    Парсит каждую строку, проверяет корректность формата и логическую последовательность диапазонов.
    Если данные валидны, сохраняет их в context.user_data["gradmark_data"] и переходит к выбору группы.
    В случае ошибки – отправляет одно из случайных сообщений из ERROR_MESSAGES_GRADMARK_FORMAT.
    """
    user_input = update.message.text.strip()
    lines = [line for line in user_input.splitlines() if line.strip()]
    num_ranges = len(lines)

    # Проверка количества диапазонов (от 1 до 100)
    if not (1 <= num_ranges <= 100) :
        await update.message.reply_text("Количество диапазонов должно быть от 1 до 100.")
        return await _gradmark_retry_or_abort(update, context)

    gradmark_data = []  # Список диапазонов в виде словарей: {"start": int, "end": int или None, "markup": float}

    # Регулярные шаблоны:
    # Закрытый диапазон: например "0-10000 1000" или "10.001-50.000 1.500"
    closed_pattern = r'^\s*(\d{1,3}(?:[.]?\d{3})*)-(\d{1,3}(?:[.]?\d{3})*)\s+(\d+(?:[.,]\d+)?%?)\s*$'
    # Открытый диапазон: например "100001+ 7500"
    open_pattern = r'^\s*(\d{1,3}(?:[.]?\d{3})*)\+\s+(\d+(?:[.,]\d+)?%?)\s*$'

    gradmark_type = context.user_data.get("gradmark_type", "fixed")

    for idx, line in enumerate(lines) :
        m_closed = re.match(closed_pattern, line)
        m_open = re.match(open_pattern, line)
        if m_closed :
            start_str, end_str, markup_str = m_closed.groups()
            # Убираем точки для группировки тысяч
            start_val = int(start_str.replace(".", ""))
            end_val = int(end_str.replace(".", ""))
        elif m_open :
            start_str, markup_str = m_open.groups()
            start_val = int(start_str.replace(".", ""))
            end_val = None  # Открытый диапазон
        else :
            return await _gradmark_retry_or_abort(update, context)

        # Обработка markup-значения: для фиксированной наценки не должно быть знака '%'
        markup_str = markup_str.replace(" ", "")
        if gradmark_type == "fixed" :
            if "%" in markup_str :
                return await _gradmark_retry_or_abort(update, context)
        # Для процентной наценки знак '%' может присутствовать, его убираем
        markup_str = markup_str.replace("%", "").replace(",", ".")
        try :
            markup_value = float(markup_str)
        except ValueError :
            return await _gradmark_retry_or_abort(update, context)

        gradmark_data.append({
            "start" : start_val,
            "end" : end_val,  # None для открытого диапазона
            "markup" : markup_value
        })

    # Проверка последовательности диапазонов:
    # Первый диапазон должен начинаться с 0
    if gradmark_data[0]["start"] != 0 :
        return await _gradmark_retry_or_abort(update, context)

    # Если диапазон закрытый, то следующий должен начинаться с (предыдущее end + 1)
    for i in range(1, len(gradmark_data)) :
        prev = gradmark_data[i - 1]
        curr = gradmark_data[i]
        if prev["end"] is None :
            # Если предыдущий диапазон уже открытый, то лишние диапазоны недопустимы
            return await _gradmark_retry_or_abort(update, context)
        expected_start = prev["end"] + 1
        if curr["start"] != expected_start :
            return await _gradmark_retry_or_abort(update, context)

    # Последний диапазон обязательно должен быть открытым (с плюсом)
    if gradmark_data[-1]["end"] is not None :
        return await _gradmark_retry_or_abort(update, context)

    # Если всё успешно, сохраняем градацию
    context.user_data["gradmark_data"] = gradmark_data
    # Переход к выбору группы (используем существующий механизм выбора группы)
    # Отправляем кнопки: "Все группы" и список групп из LIST_BRAND_GROUP
    keyboard = []
    keyboard.append([InlineKeyboardButton("Все группы", callback_data="group_ALL")])
    for group_name in LIST_BRAND_GROUP.keys() :
        callback_data = f"group_{group_name}"
        keyboard.append([InlineKeyboardButton(group_name, callback_data=callback_data)])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите группу:", reply_markup=reply_markup)
    return CHOOSE_GROUP

async def _gradmark_retry_or_abort(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int :
    """
    Вспомогательная функция для обработки неверного формата ввода градации.
    Увеличивает счетчик попыток и либо просит ввести снова, либо завершает сценарий.
    """
    attempts = context.user_data.get("gradmark_attempts", 0) + 1
    context.user_data["gradmark_attempts"] = attempts
    error_msg = random.choice(ERROR_MESSAGES_GRADMARK_FORMAT)
    if attempts >= 3 :
        await update.message.reply_text(f"{error_msg}\nСлишком много ошибок. Процесс прерван.")
        return ConversationHandler.END
    else :
        await update.message.reply_text(error_msg)
        return ENTER_GRADATION

async def gradmark_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка выбора формата вывода (CSV/Excel/Сообщение) для gradmark.
    Если выбран CSV или Excel, сначала загружаем данные и спрашиваем, в каком формате выводить таблицу.
    """
    query = update.callback_query
    choice = query.data
    await query.answer()

    context.user_data["gradmark_chosen_format"] = choice

    if choice in [FORMAT_CSV, FORMAT_EXCEL]:
        # Загружаем товары из базы и формируем итоговый список
        user_id = update.effective_user.id
        pseudo_data = load_pseudo_user_data(user_id)
        group_name = context.user_data.get("selected_group", "ALL")
        selected_brands = context.user_data.get("selected_brands", [])
        if group_name == "ALL":
            final_product_list = build_final_product_list_for_all(pseudo_data)
        else:
            final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)
        context.user_data["gradmark_final_product_list"] = final_product_list

        # Спрашиваем, в каком формате выводить таблицу
        keyboard = [
            [
                InlineKeyboardButton("Группированный формат", callback_data=GRADMARK_STRUCTURE_GROUPED),
                InlineKeyboardButton("Плоский формат", callback_data=GRADMARK_STRUCTURE_FLAT),
            ]
        ]
        await query.edit_message_text(
            text="Выберите структуру данных таблицы:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_GRADMARK_STRUCTURE
    else:
        # Если выбран формат "Сообщение", сразу переходим к вопросу о комментариях
        keyboard = [
            [
                InlineKeyboardButton("Да", callback_data=MSG_COMMENTS_YES),
                InlineKeyboardButton("Нет", callback_data=MSG_COMMENTS_NO),
            ]
        ]
        await query.edit_message_text(
            text="Учитывать комментарии?",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_COMMENTS

async def gradmark_structure_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка выбора структуры таблицы для gradmark.
    Сохраняем выбор и переходим к вопросу об учёте комментариев.
    """
    query = update.callback_query
    structure_choice = query.data  # GRADMARK_STRUCTURE_GROUPED или GRADMARK_STRUCTURE_FLAT
    await query.answer()

    if structure_choice == GRADMARK_STRUCTURE_GROUPED:
        context.user_data["gradmark_structure"] = "grouped"
    else:
        context.user_data["gradmark_structure"] = "flat"

    keyboard = [
        [
            InlineKeyboardButton("Да", callback_data=MSG_COMMENTS_YES),
            InlineKeyboardButton("Нет", callback_data=MSG_COMMENTS_NO),
        ]
    ]
    await query.edit_message_text(
        text="Учитывать комментарии?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ASK_COMMENTS

async def gradmark_comments_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Обработка ответа на вопрос об учёте комментариев для gradmark.
    Вызывает функцию генерации и отправки файла или сообщения с применением градационной наценки.
    """
    query = update.callback_query
    choice = query.data  # MSG_COMMENTS_YES или MSG_COMMENTS_NO
    await query.answer()

    include_comments = (choice == MSG_COMMENTS_YES)
    chosen_format = context.user_data.get("gradmark_chosen_format")

    if chosen_format == FORMAT_CSV:
        if context.user_data.get("gradmark_structure", "grouped") == "flat":
            await query.edit_message_text("Генерируем CSV с градационной наценкой (Плоский формат)...")
            await send_csv_flat_with_gradmark(query, context, include_comments=include_comments)
        else:
            await query.edit_message_text("Генерируем CSV с градационной наценкой...")
            await send_csv_with_gradmark(query, context, include_comments=include_comments)
    elif chosen_format == FORMAT_EXCEL:
        if context.user_data.get("gradmark_structure", "grouped") == "flat":
            await query.edit_message_text("Генерируем Excel с градационной наценкой (Плоский формат)...")
            await send_excel_flat_with_gradmark(query, context, include_comments=include_comments)
        else:
            await query.edit_message_text("Генерируем Excel с градационной наценкой...")
            await send_excel_with_gradmark(query, context, include_comments=include_comments)
    elif chosen_format == FORMAT_MSG:
        await query.edit_message_text("Формируем сообщение с градационной наценкой...")
        await send_message_with_gradmark(query, context, include_comments=include_comments)
    else:
        await query.edit_message_text("Неизвестный формат.")

    return ConversationHandler.END

# ------------------------------ Функции для применения градационной наценки ------------------------------

def apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data) :
    """
    Применяет градационную наценку к ценам товаров.
    Для каждого товара определяется, в какой диапазон попадает его цена, и к цене применяется соответствующая наценка.

    gradmark_type: "fixed" или "percent"
    gradmark_data: список диапазонов вида:
        [{"start": int, "end": int или None, "markup": float}, ...]
    """
    if not final_product_list :
        return

    for model_group, product_dict in final_product_list.items() :
        for product_name, entries in product_dict.items() :
            for entry in entries :
                orig_price = entry["price"]
                # Ищем подходящий диапазон
                applicable_markup = None
                for rng in gradmark_data :
                    start = rng["start"]
                    end = rng["end"]
                    if end is not None :
                        if start <= orig_price <= end :
                            applicable_markup = rng["markup"]
                            break
                    else :
                        # Открытый диапазон (N+)
                        if orig_price >= start :
                            applicable_markup = rng["markup"]
                            break
                # Если подходящий диапазон найден, применяем наценку
                if applicable_markup is not None :
                    if gradmark_type == "fixed" :
                        new_price = orig_price + applicable_markup
                    else :  # percent
                        new_price = orig_price + (orig_price * (applicable_markup / 100.0))
                        new_price = round_up_to_nearest_100(new_price)
                    entry["price"] = int(new_price)
    # Если по каким-то причинам товар не попал ни в один диапазон – оставляем цену без изменений.

# ------------------------------ Функции отправки файлов/сообщений с градационной наценкой ------------------------------

async def send_csv_with_gradmark(query, context, include_comments=False) :
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    selected_brands = context.user_data.get("selected_brands", [])
    gradmark_type = context.user_data.get("gradmark_type", "fixed")
    gradmark_data = context.user_data.get("gradmark_data", [])

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)

    apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data)

    if not include_comments :
        remove_commented_products(final_product_list)

    await send_csv(query.message, context, final_product_list, file_prefix="gradmark_")

async def send_csv_flat_with_gradmark(query, context, include_comments=False) :
    """
    Генерирует CSV-файл для /gradmark в плоском формате с градационной наценкой.
    """
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    selected_brands = context.user_data.get("selected_brands", [])
    gradmark_type = context.user_data.get("gradmark_type", "fixed")
    gradmark_data = context.user_data.get("gradmark_data", [])
    pseudo_data = load_pseudo_user_data(user_id)

    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)

    apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data)
    if not include_comments :
        remove_commented_products(final_product_list)

    await send_csv_flat(query.message, context, final_product_list, file_prefix="gradmark_")

async def send_excel_with_gradmark(query, context, include_comments=False):
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    selected_brands = context.user_data.get("selected_brands", [])
    gradmark_type = context.user_data.get("gradmark_type", "fixed")
    gradmark_data = context.user_data.get("gradmark_data", [])

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)

    apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data)

    if not include_comments:
        remove_commented_products(final_product_list)

    await send_excel(query.message, context, final_product_list, file_prefix="gradmark_")

async def send_excel_flat_with_gradmark(query, context, include_comments=False) :
    """
    Генерирует Excel-файл для /gradmark в плоском формате с градационной наценкой.
    """
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    selected_brands = context.user_data.get("selected_brands", [])
    gradmark_type = context.user_data.get("gradmark_type", "fixed")
    gradmark_data = context.user_data.get("gradmark_data", [])
    pseudo_data = load_pseudo_user_data(user_id)

    if group_name == "ALL" :
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else :
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)

    apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data)
    if not include_comments :
        remove_commented_products(final_product_list)

    await send_excel_flat(query.message, context, final_product_list, file_prefix="gradmark_")

async def send_message_with_gradmark(query, context, include_comments=False):
    user_id = query.message.chat_id
    group_name = context.user_data.get("selected_group", "ALL")
    selected_brands = context.user_data.get("selected_brands", [])
    gradmark_type = context.user_data.get("gradmark_type", "fixed")
    gradmark_data = context.user_data.get("gradmark_data", [])

    pseudo_data = load_pseudo_user_data(user_id)
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands(pseudo_data, selected_brands)

    apply_gradmark_to_final_list(final_product_list, gradmark_type, gradmark_data)

    messages = generate_text_messages(
        final_product_list,
        PRODUCT_LIBRARY,
        include_comments=include_comments
    )
    for msg in messages:
        await query.message.reply_text(msg, parse_mode="HTML")

# ------------------------------ ConversationHandler для /gradmark ------------------------------

def get_gradmark_conversation_handler():
    return ConversationHandler(
        entry_points=[CommandHandler("gradmark", gradmark_command)],
        states={
            gradmark_TYPE: [
                CallbackQueryHandler(gradmark_type_callback, pattern=f"^{gradmark_FIXED}|{gradmark_PERCENT}$")
            ],
            ENTER_GRADATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, gradmark_gradation_handler)
            ],
            CHOOSE_GROUP: [
                CallbackQueryHandler(markup_group_callback, pattern=r"^group_")
            ],
            CHOOSE_CATEGORY: [
                CallbackQueryHandler(markup_category_callback, pattern=r"^cat_")
            ],
            CHOOSE_FORMAT: [
                CallbackQueryHandler(gradmark_format_callback, pattern=f"^{FORMAT_CSV}|{FORMAT_EXCEL}|{FORMAT_MSG}$")
            ],
            ASK_GRADMARK_STRUCTURE: [
                CallbackQueryHandler(gradmark_structure_callback, pattern=f"^{GRADMARK_STRUCTURE_GROUPED}|{GRADMARK_STRUCTURE_FLAT}$")
            ],
            ASK_COMMENTS: [
                CallbackQueryHandler(gradmark_comments_callback, pattern=f"^{MSG_COMMENTS_YES}|{MSG_COMMENTS_NO}$")
            ],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
    )

# ------------------------------ ФУНКЦИИ ДЛЯ ФОРМИРОВАНИЯ LIST ------------------------------

# Для команды /list
LIST_GROUP = 5       # Выбор группы
LIST_CATEGORY = 6    # Выбор категории
LIST_FORMAT = 7      # Выбор формата (CSV/Excel/Сообщение)

# Новые константы для выбора структуры таблицы в /list
LIST_STRUCTURE_GROUPED = "list_structure_grouped"
LIST_STRUCTURE_FLAT = "list_structure_flat"

# Добавляем новое состояние для выбора структуры:
ASK_LIST_STRUCTURE = 8


def build_full_product_list(user_data: dict) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Группируем товары по (model_group, product_name),
    удаляем дубли и возвращаем структуру:
         final_product_list[model_group][product_name] -> List[offers]
    При этом в каждом оффере сохраняем поля: "price", "country", "supplier", "comment" и "brand".
    """
    logger = logging.getLogger(__name__)
    products_data = user_data.get("products", [])
    grouped_products: Dict[Tuple[str, str], Set[Tuple[Any, str, str, str, str]]] = defaultdict(set)

    for product in products_data:
        model_group = product.get("model_group")
        product_name = product.get("product_name")
        # Предполагаем, что в данных присутствует поле "brand"
        brand = product.get("brand", "Не указан")
        country = product.get("country", "Не указано")
        price = product.get("price")
        supplier = product.get("supplier", "Не указан")
        comment = product.get("comment", "Без комментариев")

        if not model_group or not product_name or price is None:
            logger.info(f"[build_full_product_list] Пропущен товар из-за отсутствия данных: {product}")
            continue

        key = (model_group, product_name)
        offer_tuple = (price, country, supplier, comment, brand)
        grouped_products[key].add(offer_tuple)

    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(dict)
    for (model_group, product_name), offers in grouped_products.items():
        final_product_list[model_group].setdefault(product_name, [])
        # Сортируем офферы по цене (по возрастанию)
        sorted_offers = sorted(offers, key=lambda x: x[0])
        for price, country, supplier, comment, brand in sorted_offers:
            final_product_list[model_group][product_name].append({
                "price": price,
                "country": country,
                "supplier": supplier,
                "comment": comment,
                "brand": brand
            })
    return final_product_list

def filter_products_by_brand(full_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
                             accepted_brands: List[str]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Фильтрует итоговый список товаров full_list так, чтобы оставить только те офферы,
    у которых значение поля "brand" содержится в списке accepted_brands.
    """
    filtered = {}
    for mg, prod_dict in full_list.items():
        for product_name, offers in prod_dict.items():
            filtered_offers = [offer for offer in offers if offer.get("brand") in accepted_brands]
            if filtered_offers:
                if mg not in filtered:
                    filtered[mg] = {}
                filtered[mg][product_name] = filtered_offers
    return filtered

def build_final_product_list_for_brand_list(user_data: dict, selected_brand: str) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Формирует итоговый список товаров для одного выбранного подбренда.
    Если selected_brand == "ALL" – возвращаем весь список.
    Иначе фильтруем товары так, чтобы оставались только те, у которых поле "brand" равно selected_brand.
    """
    full_list = build_full_product_list(user_data)
    if selected_brand == "ALL":
        return full_list
    return filter_products_by_brand(full_list, [selected_brand])

def build_final_product_list_for_multiple_brands_list(user_data: dict, brand_list: List[str]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Объединяет результаты для каждого выбранного подбренда (например, для группы "Anrdoid"
    это может быть ["Samsung Galaxy", "Xiaomi", "OnePlus"]).
    """
    if not brand_list:
        return {}
    full_list = build_full_product_list(user_data)
    return filter_products_by_brand(full_list, brand_list)

def build_final_product_list_for_all_list(user_data: dict) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Возвращает полный список товаров (аналог выбора "Все группы").
    """
    return build_full_product_list(user_data)

async def list_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id

    # Проверяем, есть ли товары у пользователя (в БД)
    rows = get_all_products(user_id)
    if not rows:
        await update.message.reply_text("Ваш список товаров пуст. Сначала отправьте товары.")
        return ConversationHandler.END

    # Раз товары есть — продолжаем
    keyboard = []
    keyboard.append([InlineKeyboardButton("Все группы", callback_data="group_ALL")])
    for group_name in LIST_BRAND_GROUP.keys():
        callback_data = f"group_{group_name}"
        keyboard.append([InlineKeyboardButton(group_name, callback_data=callback_data)])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите группу:", reply_markup=reply_markup)

    return LIST_GROUP

async def list_group_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Если group_ALL => сразу к LIST_FORMAT (выбор формата)
    Иначе => показываем "Все категории" + sub-бренды из LIST_BRAND_GROUP[group].
    """
    query = update.callback_query
    data = query.data  # "group_ALL" / "group_Apple" / ...
    await query.answer()

    if data == "group_ALL":
        context.user_data["list_selected_group"] = "ALL"
        # Переходим сразу к выбору формата
        keyboard = [
            [
                InlineKeyboardButton("CSV", callback_data="list_csv"),
                InlineKeyboardButton("Excel", callback_data="list_excel"),
                InlineKeyboardButton("Сообщение", callback_data="list_msg")
            ]
        ]
        await query.edit_message_text("Выберите формат вывода списка товаров:", reply_markup=InlineKeyboardMarkup(keyboard))
        return LIST_FORMAT

    # Иначе конкретная группа
    group_name = data.replace("group_", "")
    context.user_data["list_selected_group"] = group_name

    # Показываем "Все категории" + sub-бренды
    sub_brands = LIST_BRAND_GROUP.get(group_name, [])

    keyboard = []
    keyboard.append([InlineKeyboardButton("Все категории", callback_data="cat_ALL")])
    for sb in sub_brands:
        cb = f"cat_{sb}"
        keyboard.append([InlineKeyboardButton(sb, callback_data=cb)])

    await query.edit_message_text("Выберите категорию:", reply_markup=InlineKeyboardMarkup(keyboard))
    return LIST_CATEGORY

async def list_category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    Если cat_ALL => выбираем все под-бренды группы
    Иначе => один конкретный sub-brand
    Переходим к выбору формата (CSV/Excel/Сообщение).
    """
    query = update.callback_query
    data = query.data  # cat_ALL / cat_Apple iPhone / ...
    await query.answer()

    group_name = context.user_data.get("list_selected_group", "ALL")
    if data == "cat_ALL":
        # все sub-бренды
        sub_brands = LIST_BRAND_GROUP.get(group_name, [])
        context.user_data["list_selected_sub_brands"] = sub_brands
    else:
        brand_str = data.replace("cat_", "")
        context.user_data["list_selected_sub_brands"] = [brand_str]

    # Показываем формат
    keyboard = [
        [
            InlineKeyboardButton("CSV", callback_data="list_csv"),
            InlineKeyboardButton("Excel", callback_data="list_excel"),
            InlineKeyboardButton("Сообщение", callback_data="list_msg")
        ]
    ]
    await query.edit_message_text("Выберите формат вывода списка товаров:", reply_markup=InlineKeyboardMarkup(keyboard))
    return LIST_FORMAT

async def list_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    choice = query.data  # "list_csv", "list_excel" или "list_msg"
    await query.answer()

    user_id = query.message.chat_id
    group_name = context.user_data.get("list_selected_group", "ALL")
    sub_brands = context.user_data.get("list_selected_sub_brands", [])

    # Загружаем товары из базы
    pseudo_data = load_pseudo_user_data(user_id)

    # Составляем итоговый список товаров
    if group_name == "ALL":
        final_product_list = build_final_product_list_for_all_list(pseudo_data)
    else:
        final_product_list = build_final_product_list_for_multiple_brands_list(pseudo_data, sub_brands)

    if not final_product_list:
        await query.edit_message_text("Ничего не найдено для этой группы/категории.")
        return ConversationHandler.END

    # Сохраняем итоговый список и выбранный формат
    context.user_data["list_final_product_list"] = final_product_list
    context.user_data["list_chosen_format"] = choice

    if choice in ["list_csv", "list_excel"]:
        # Запрашиваем структуру таблицы
        keyboard = [
            [
                InlineKeyboardButton("Группированный формат", callback_data=LIST_STRUCTURE_GROUPED),
                InlineKeyboardButton("Плоский формат", callback_data=LIST_STRUCTURE_FLAT),
            ]
        ]
        await query.edit_message_text(
            text="Выберите структуру данных таблицы:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ASK_LIST_STRUCTURE
    elif choice == "list_msg":
        await query.edit_message_text("Отправляем список товаров...")
        messages = generate_full_text_messages(final_product_list, group_name)
        for msg in messages:
            await query.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return ConversationHandler.END
    else:
        await query.edit_message_text("Неизвестный формат.")
        return ConversationHandler.END

async def list_structure_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    structure_choice = query.data  # LIST_STRUCTURE_GROUPED или LIST_STRUCTURE_FLAT
    await query.answer()

    # Сохраняем выбор структуры
    context.user_data["list_structure"] = "grouped" if structure_choice == LIST_STRUCTURE_GROUPED else "flat"

    # Определяем, какой формат выбран (CSV или Excel)
    chosen_format = context.user_data.get("list_chosen_format")
    final_product_list = context.user_data.get("list_final_product_list")
    group_name = context.user_data.get("list_selected_group", "ALL")

    if chosen_format == "list_csv":
        if context.user_data["list_structure"] == "flat":
            await send_list_csv_flat(query.message, context, final_product_list, group_name)
        else:
            await send_list_csv(query.message, context, final_product_list, group_name)
    elif chosen_format == "list_excel":
        if context.user_data["list_structure"] == "flat":
            await send_list_excel_flat(query.message, context, final_product_list, group_name)
        else:
            await send_list_excel(query.message, context, final_product_list, group_name)
    return ConversationHandler.END

# --------------------- CSV ---------------------

async def send_list_csv(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    selected_brand: str
):
    """
    Формирует CSV-файл.
    Если выбран "ALL", идём по PRODUCT_LIBRARY,
    иначе – для конкретного бренда обходим model_group из final_product_list.
    """
    logger = logging.getLogger(__name__)
    logger.info("[send_list_csv] Формируем CSV...")

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"full_product_list_{current_date}.csv"

    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий"]
            writer.writerow(header)

            if selected_brand == "ALL":
                # Обходим по группам из PRODUCT_LIBRARY
                brand_list = list(PRODUCT_LIBRARY.keys())
                for brand in brand_list:
                    # Для каждого бренда обходим его model_group (ключи PRODUCT_LIBRARY[brand])
                    for model_group in PRODUCT_LIBRARY.get(brand, {}):
                        if model_group not in final_product_list:
                            continue
                        writer.writerow([model_group, "Цена", "Страны", "Поставщик", "Комментарий"])

                        # "Разворачиваем" словарь: для каждого product_name и каждой записи (entry)
                        flattened_products = [
                            (product_name, entry)
                            for product_name, entries in final_product_list[model_group].items()
                            for entry in entries
                        ]

                        # Многокритериальная сортировка:
                        # 1. По get_sort_key(brand, model_group, product_name)
                        # 2. По цене (entry["price"]) по возрастанию, если product_name одинаковые
                        sorted_products = sorted(
                            flattened_products,
                            key=lambda x : (
                                get_sort_key(brand, model_group, x[0]),
                                x[1]["price"]
                            )
                        )

                        for product_name, entry in sorted_products :
                            writer.writerow([
                                product_name,
                                entry["price"],
                                entry["country"],
                                entry["supplier"],
                                entry["comment"]
                            ])
                        writer.writerow([])

            else:
                # Для конкретного бренда обходим по всем model_group из final_product_list
                for model_group in sorted(final_product_list.keys()):
                    writer.writerow([model_group, "Цена", "Страны", "Поставщик", "Комментарий"])
                    products = final_product_list[model_group]
                    # Используем выбранный бренд для сортировки
                    sorted_items = sorted(
                        products.items(),
                        key=lambda x: get_sort_key(selected_brand, model_group, x[0])
                    )
                    for product_name, entries in sorted_items:
                        sorted_entries = sorted(entries, key=lambda x: x["price"])
                        for entry in sorted_entries:
                            writer.writerow([
                                product_name,
                                entry["price"],
                                entry["country"],
                                entry["supplier"],
                                entry["comment"]
                            ])
                    writer.writerow([])

        with open(filename, "rb") as f:
            await message.reply_document(document=f, filename=filename, caption="Полный список (CSV)!")
    except Exception as e:
        logger.error(f"Ошибка при формировании CSV: {e}", exc_info=True)
        await message.reply_text("Произошла ошибка при формировании CSV.")
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except Exception as e:
                logger.warning(f"Ошибка удаления файла {filename}: {e}")

async def send_list_csv_flat(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    selected_brand: str
):
    """
    Формирует CSV-файл для /list в плоском формате.
    Добавляются две дополнительные колонки: "Категория" (model_group) и "Группа" (brand).
    """
    logger = logging.getLogger(__name__)
    logger.info("[send_list_csv_flat] Формируем CSV (Плоский формат)...")

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"full_product_list_{current_date}.csv"

    try:
        with open(filename, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            # Новый заголовок с дополнительными колонками
            header = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"]
            writer.writerow(header)

            if selected_brand == "ALL":
                brand_list = list(PRODUCT_LIBRARY.keys())
                for brand in brand_list:
                    for model_group in PRODUCT_LIBRARY.get(brand, {}):
                        if model_group not in final_product_list:
                            continue
                        writer.writerow([model_group, "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"])
                        # "Разворачиваем" словарь
                        flattened_products = [
                            (product_name, entry)
                            for product_name, entries in final_product_list[model_group].items()
                            for entry in entries
                        ]
                        # Сортировка по get_sort_key и цене
                        sorted_products = sorted(
                            flattened_products,
                            key=lambda x: (get_sort_key(brand, model_group, x[0]), x[1]["price"])
                        )
                        for product_name, entry in sorted_products:
                            writer.writerow([
                                product_name,
                                entry["price"],
                                entry["country"],
                                entry["supplier"],
                                entry["comment"],
                                model_group,  # Категория
                                brand         # Группа
                            ])
                        writer.writerow([])
            else:
                for model_group in sorted(final_product_list.keys()):
                    writer.writerow([model_group, "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"])
                    products = final_product_list[model_group]
                    sorted_items = sorted(
                        products.items(),
                        key=lambda x: get_sort_key(selected_brand, model_group, x[0])
                    )
                    for product_name, entries in sorted_items:
                        sorted_entries = sorted(entries, key=lambda x: x["price"])
                        for entry in sorted_entries:
                            writer.writerow([
                                product_name,
                                entry["price"],
                                entry["country"],
                                entry["supplier"],
                                entry["comment"],
                                model_group,  # Категория
                                selected_brand  # Группа
                            ])
                    writer.writerow([])

        with open(filename, "rb") as f:
            await message.reply_document(document=f, filename=filename, caption="Полный список (CSV, Плоский формат)!")
    except Exception as e:
        logger.error(f"Ошибка при формировании CSV (плоский формат): {e}", exc_info=True)
        await message.reply_text("Произошла ошибка при формировании CSV (плоский формат).")
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except Exception as e:
                logger.warning(f"Ошибка удаления файла {filename}: {e}")

# --------------------- EXCEL ---------------------

async def send_list_excel(
    message: Message,
    context: ContextTypes.DEFAULT_TYPE,
    final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
    selected_brand: str
):
    """
    Формирует Excel-файл.
    Если выбран "ALL", идём по PRODUCT_LIBRARY,
    иначе – для конкретного бренда обходим model_group из final_product_list.
    """
    logger = logging.getLogger(__name__)
    logger.info("[send_list_excel] Формируем Excel...")

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"full_product_list_{current_date}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Полный Список Товаров"

        headers = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий"]
        ws.append(headers)

        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        if selected_brand == "ALL":
            brand_list = list(PRODUCT_LIBRARY.keys())
            for brand in brand_list:
                for model_group in PRODUCT_LIBRARY.get(brand, {}):
                    if model_group not in final_product_list:
                        continue
                    ws.append([model_group, "Цена", "Страны", "Поставщик", "Комментарий"])
                    for col in range(1, len(headers) + 1) :
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                               fill_type="solid")

                    products = final_product_list[model_group]

                    # "Разворачиваем" словарь: для каждого product_name и каждой записи (entry)
                    flattened_products = [
                        (product_name, entry)
                        for product_name, entries in products.items()
                        for entry in entries
                    ]

                    # Многокритериальная сортировка:
                    # 1. По результату get_sort_key(brand, model_group, product_name)
                    # 2. По цене (entry["price"]) по возрастанию, если product_name совпадают
                    sorted_products = sorted(
                        flattened_products,
                        key=lambda x : (
                            get_sort_key(brand, model_group, x[0]),
                            x[1]["price"]
                        )
                    )

                    # Записываем отсортированные данные в Excel
                    for product_name, entry in sorted_products :
                        ws.append([
                            product_name,
                            entry["price"],
                            entry["country"],
                            entry["supplier"],
                            entry["comment"]
                        ])

                    ws.append([])  # Пустая строка после группы
        else:
            # Для конкретного бренда обходим все model_group из final_product_list
            for model_group in sorted(final_product_list.keys()):
                ws.append([model_group, "Цена", "Страны", "Поставщик", "Комментарий"])
                for col in range(1, len(headers)+1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                products = final_product_list[model_group]
                sorted_items = sorted(
                    products.items(),
                    key=lambda x: get_sort_key(selected_brand, model_group, x[0])
                )
                for product_name, entries in sorted_items:
                    sorted_entries = sorted(entries, key=lambda x: x["price"])
                    for entry in sorted_entries:
                        ws.append([
                            product_name,
                            entry["price"],
                            entry["country"],
                            entry["supplier"],
                            entry["comment"]
                        ])
                ws.append([])

        wb.save(filename)
        with open(filename, "rb") as f:
            await message.reply_document(document=f, filename=filename, caption="Полный список (Excel)!")
    except Exception as e:
        logger.error(f"Ошибка при формировании Excel: {e}", exc_info=True)
        await message.reply_text("Произошла ошибка при формировании Excel.")
    finally:
        if os.path.exists(filename):
            try:
                os.remove(filename)
            except Exception as e:
                logger.warning(f"Ошибка удаления файла {filename}: {e}")


async def send_list_excel_flat(
        message: Message,
        context: ContextTypes.DEFAULT_TYPE,
        final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
        selected_brand: str
) :
    """
    Формирует Excel-файл для /list в плоском формате с дополнительными колонками:
    "Категория" (model_group) и "Группа" (brand).
    """
    logger = logging.getLogger(__name__)
    logger.info("[send_list_excel_flat] Формируем Excel (Плоский формат)...")

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"full_product_list_{current_date}.xlsx"

    try :
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Полный Список Товаров (Плоский)"

        headers = ["Наименование", "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"]
        ws.append(headers)
        bold_font = Font(bold=True)
        for cell in ws[1] :
            cell.font = bold_font

        if selected_brand == "ALL" :
            brand_list = list(PRODUCT_LIBRARY.keys())
            for brand in brand_list :
                for model_group in PRODUCT_LIBRARY.get(brand, {}) :
                    if model_group not in final_product_list :
                        continue
                    ws.append([model_group, "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"])
                    for col in range(1, len(headers) + 1) :
                        ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                               fill_type="solid")

                    products = final_product_list[model_group]
                    flattened_products = [
                        (product_name, entry)
                        for product_name, entries in products.items()
                        for entry in entries
                    ]
                    sorted_products = sorted(
                        flattened_products,
                        key=lambda x : (get_sort_key(brand, model_group, x[0]), x[1]["price"])
                    )
                    for product_name, entry in sorted_products :
                        ws.append([
                            product_name,
                            entry["price"],
                            entry["country"],
                            entry["supplier"],
                            entry["comment"],
                            model_group,  # Категория
                            brand  # Группа
                        ])
                    ws.append([])
        else :
            for model_group in sorted(final_product_list.keys()) :
                ws.append([model_group, "Цена", "Страны", "Поставщик", "Комментарий", "Категория", "Группа"])
                for col in range(1, len(headers) + 1) :
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                           fill_type="solid")
                products = final_product_list[model_group]
                sorted_items = sorted(
                    products.items(),
                    key=lambda x : get_sort_key(selected_brand, model_group, x[0])
                )
                for product_name, entries in sorted_items :
                    sorted_entries = sorted(entries, key=lambda x : x["price"])
                    for entry in sorted_entries :
                        ws.append([
                            product_name,
                            entry["price"],
                            entry["country"],
                            entry["supplier"],
                            entry["comment"],
                            model_group,  # Категория
                            selected_brand  # Группа
                        ])
                ws.append([])

        wb.save(filename)
        with open(filename, "rb") as f :
            await message.reply_document(document=f, filename=filename,
                                         caption="Полный список (Excel, Плоский формат)!")
    except Exception as e :
        logger.error(f"Ошибка при формировании Excel (плоский формат): {e}", exc_info=True)
        await message.reply_text("Произошла ошибка при формировании Excel (плоский формат).")
    finally :
        if os.path.exists(filename) :
            try :
                os.remove(filename)
            except Exception as e :
                logger.warning(f"Ошибка удаления файла {filename}: {e}")

# --------------------- ТЕКСТОВОЕ СООБЩЕНИЕ ---------------------

def generate_full_text_messages(
        final_product_list: Dict[str, Dict[str, List[Dict[str, Any]]]],
        selected_brand: str
) -> List[str]:
    """
    Формирует текстовый список (HTML-разметка).

    Если выбран конкретный бренд (например, "Samsung Galaxy" или "Фены для волос Dyson"):
      - Сначала выводится заголовок этого бренда (жирным с курсивом, верхний регистр).
      - Затем для каждой модели (model_group) для данного бренда выводится её название (жирным),
        а далее для каждого товара (product_name) – его строка (жирным с курсивом), с офферами,
        отсортированными с использованием get_sort_key.

    Если выбран "ALL":
      - Товары группируются по брендам (используя значение поля "brand" у офферов) в том же порядке,
        как они идут в PRODUCT_LIBRARY.
      - Для каждого бренда выводится заголовок (жирным с курсивом, верхний регистр),
        затем для каждой модели – её название (жирным) и товары.
    """
    logger = logging.getLogger(__name__)
    logger.info("[generate_full_text_messages] Формируем текстовый список для /list...")

    lines: List[str] = []

    if selected_brand == "ALL":
        # Сгруппируем final_product_list по брендам.
        # Для каждого model_group берем бренд из первой записи (предполагается, что все офферы в группе имеют одинаковый brand)
        grouped_by_brand: Dict[str, Dict[str, Dict[str, List[Dict[str, Any]]]]] = {}
        for model_group, products in final_product_list.items():
            # Ищем бренд из одной из записей
            brand_val = None
            for prod_name, offers in products.items():
                if offers:
                    brand_val = offers[0].get("brand", "").strip()
                    break
            if not brand_val:
                continue
            if brand_val not in grouped_by_brand:
                grouped_by_brand[brand_val] = {}
            grouped_by_brand[brand_val][model_group] = products

        # Перебираем бренды в том же порядке, что и в PRODUCT_LIBRARY
        for brand in PRODUCT_LIBRARY.keys():
            # Если для этого бренда нет товаров, пропускаем
            if brand not in grouped_by_brand:
                continue
            # Заголовок бренда
            lines.append(f"<b><i>{brand.upper()}</i></b>")
            lines.append("")
            # Определяем порядок моделей для этого бренда.
            # Если для high-level бренда есть список в LIST_BRAND_GROUP, используем его порядок,
            # иначе – порядок по ключам сгруппированных моделей.
            if brand in LIST_BRAND_GROUP:
                desired_model_order = LIST_BRAND_GROUP[brand]
                # Отфильтруем desired_model_order, оставив только те model_group, которые есть в grouped_by_brand[brand]
                model_order = [m for m in desired_model_order if m in grouped_by_brand[brand]]
                # Дополнительно добавим оставшиеся модели (отсортированные по алфавиту)
                remaining = sorted(set(grouped_by_brand[brand].keys()) - set(model_order))
                model_order.extend(remaining)
            else:
                model_order = sorted(grouped_by_brand[brand].keys())

            for model_group in model_order:
                lines.append(f"<b>{model_group}</b>")
                lines.append("")
                products = grouped_by_brand[brand][model_group]
                # Получаем словарь товаров для текущей модели
                products = final_product_list[model_group]

                # "Разворачиваем" словарь: для каждого product_name и для каждой записи (entry)
                flattened_products = [
                    (product_name, entry)
                    for product_name, entries in products.items()
                    for entry in entries
                ]

                # Многокритериальная сортировка:
                # 1. По результату get_sort_key(brand, model_group, product_name)
                # 2. По цене (entry["price"]) по возрастанию, если product_name совпадают
                sorted_products = sorted(
                    flattened_products,
                    key=lambda x : (
                        get_sort_key(brand, model_group, x[0]),
                        x[1]["price"]
                    )
                )

                # Группируем по product_name, чтобы вывести заголовок продукта один раз
                for product_name, group in itertools.groupby(sorted_products, key=lambda x : x[0]) :
                    lines.append(f"<b><i>{product_name}</i></b>:")
                    for _, entry in group :
                        price = entry["price"]
                        comment = entry["comment"]
                        country_name = entry["country"]
                        country_flag = TEXT_TO_FLAG.get(country_name, country_name)
                        supplier = entry["supplier"]
                        formatted_price = f"{price:,}".replace(",", ".")
                        lines.append(f"• {formatted_price} руб. | {country_flag} | {supplier} | {comment}")
                    lines.append("")  # Пустая строка после группы товаров

                lines.append("")

    # Удаляем лишние пустые строки в конце
    while lines and not lines[-1].strip():
        lines.pop()

    # Разбиваем на чанки, не превышающие MAX_TELEGRAM_TEXT
    messages: List[str] = []
    current_chunk = ""
    for line in lines:
        candidate_len = len(current_chunk) + len(line) + 1
        if candidate_len > MAX_TELEGRAM_TEXT:
            messages.append(current_chunk)
            current_chunk = line
        else:
            if not current_chunk:
                current_chunk = line
            else:
                current_chunk += "\n" + line
    if current_chunk:
        messages.append(current_chunk)

    logger.info(f"[generate_full_text_messages] Получено {len(messages)} чанков для отправки.")
    return messages

def get_list_conversation_handler():
    return ConversationHandler(
        entry_points=[CommandHandler("list", list_command)],
        states={
            LIST_GROUP: [CallbackQueryHandler(list_group_callback, pattern=r"^group_")],
            LIST_CATEGORY: [CallbackQueryHandler(list_category_callback, pattern=r"^cat_")],
            LIST_FORMAT: [CallbackQueryHandler(list_format_callback, pattern=r"^list_(csv|excel|msg)$")],
            ASK_LIST_STRUCTURE: [CallbackQueryHandler(list_structure_callback, pattern=r"^(list_structure_grouped|list_structure_flat)$")],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        name="list_conversation",
        persistent=False
    )

# ------------------------------ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ БОТА------------------------------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None :
    await update.message.reply_text(
        "Привет! Я бот <b><Быстрый Прайс></b>. Отправьте сообщения с товарами и используйте команды для их обработки, если нужна помощь нажмите /help.")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None :
    user_first_name = update.effective_user.first_name

    help_text = (
        "<b>💡 Справка по командам бота</b>\n\n"
        f"Привет, {user_first_name}! Рад, что ты используешь нашего бота. Ниже описаны его ключевые возможности, команды и функции, которые помогут вам работать <b>быстрее и эффективнее</b>:\n\n"

        "<b>📌 Команды:</b>\n\n"
        "<b>• /start</b> — <i>Запустить бота</i>.\n"
        "<b>• /best</b> — <i>Список товаров по лучшим ценам</i>.\n"
        "<b>• /markup</b> — <i>Список товаров с вашей наценкой</i>.\n"
        "<b>• /gradmark</b> — <i>Список товаров с градацией наценки</i>.\n"
        "<b>• /list</b> — <i>Полный список товаров</i>.\n"
        "<b>• /clear</b> — <i>Очистка списка товаров</i>.\n"
        "<b>• /my_price_list</b> — <i>Создание уникального прайс-листа</i>.\n"
        "<b>• /restart_price_list</b> — <i>Сброс настроек прайс-листа</i>.\n"
        "<b>• /cancel</b> — <i>Отмена цикла создания прайс-листа</i>.\n"
        "<b>• /currency</b> — <i>Актуальные курсы валют</i>.\n"
        "<b>• /restart</b> — <i>Сброс всех состояний в операциях</i>.\n"
        "<b>• /help</b> — <i>Получить справку по командам</i>.\n\n"

        "<b>🔍 Особенности работы бота:</b>\n\n"
        "• <i><b>✅ Гибкая настройка прайс-листа:</b></i> возможность самостоятельно настраивать столбцы, задавать градацию цен и экспортировать данные в удобном формате.\n"
        "• <i><b>✅ Интеллектуальный анализ формата:</b></i> автоматически определяет и правильно извлекает данные из многострочных и однострочных прайс-листов.\n"
        "• <i><b>✅ Учёт комментариев:</b></i> фиксирует важные детали о товаре (состояние упаковки, особенности, б/у...).\n"
        "• <i><b>✅ Анализ эмодзи:</b></i> определяет эмодзи 📦, 🚚, ✈️ и др., указывающие на статус товара.\n"
        "• <i><b>✅ Добавление поставщиков:</b></i> фиксирует продавцов из сообщений и помогает сразу понять, кто предлагает товар.\n"
        "• <i><b>✅ Распознавание стран:</b></i> распознаёт все страны мира по эмодзи и текстовым паттернам.\n"
        "• <i><b>✅ Большая библиотека товаров:</b></i> ежедневные обновления и расширение базы товаров.\n\n"

        "<b>⚙️ Как работает бот?</b>\n\n"
        "Просто <b>перешлите сообщение</b> с товарами – бот автоматически <b>извлечёт всю важную информацию</b> и добавит её в список.\n\n"
        "После этого можно использовать одну из указанных команд.\n\n"

        "<b>💡 Важно!</b> Если один и тот же товар идёт с разными странами – он будет считаться <b>разными позициями</b>.\n\n"

        "📌 Если вы не пересылаете сообщение, а отправляете его вручную и <b>хотите указать</b> поставщика, просто добавьте его в конце строки в формате: [Название поставщика].\n\n"

        "<b>📢 Цены могут меняться в течение дня!</b> Чтобы избежать неактуальных данных, <b>периодически очищайте список</b> командой /clear и отправляете новые сообщения.\n\n"

        "🚀 <b>Бот абсолютно бесплатный!</b>\n\n"
        "Сейчас мы на финальной стадии разработки и проводим <b>открытое бета-тестирование</b>. Это значит, что у вас есть <b>уникальная возможность</b> опробовать бота <b>раньше других</b>, <b>влиять на его развитие</b> и <b>получать доступ к новым функциям первым</b>.\n\n"

        "🔧 Мы активно разрабатываем <b>новые функции</b> и тестируем <b>новые возможности</b> для автоматизации работы.\n\n"

        "💬 <b>Нам важно ваше мнение!</b> Как вам бот? Что можно улучшить? Какие функции добавить? Ваши идеи и фидбек помогут нам сделать сервис лучше.\n\n"

        "📩 <b>Попробуйте прямо сейчас</b> и расскажите нам о своём опыте — <b>для нас это действительно важно!</b>\n\n"

        "Если у вас возникли <b>вопросы или предложения</b>, свяжитесь с нами @admnlab"
    )

    await update.effective_message.reply_text(
        help_text,
        parse_mode=ParseMode.HTML
    )

async def clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    # ---------------------
    # Раньше:
    #   USER_DATA[user_id] = {"products" : []}
    # теперь:
    clear_user_data(user_id)
    await update.message.reply_text("Ваш список товаров был очищен.")
    # Также удаляем настройки прайс-листа


async def currency_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Получаем данные от всех API
    cbr_success = currency_api_cbr.get_currency_rates()
    exchange_success = exchange_api.get_rates()

    # Инициализация и получение данных Binance
    binance_api = BinanceAPI()

    # Спотовые данные
    binance_spot = {
        'USDT_RUB': binance_api.get_binance_data('USDTRUB'),
        'BTC_USD': binance_api.get_binance_data('BTCUSDT'),
        'ETH_USD': binance_api.get_binance_data('ETHUSDT')
    }

    # P2P данные
    binance_p2p_buy = binance_api.get_p2p_rate(trade_type="BUY")
    binance_p2p_sell = binance_api.get_p2p_rate(trade_type="SELL")

    binance_success = any(binance_spot.values()) or (binance_p2p_buy and binance_p2p_sell)

    if not cbr_success and not exchange_success and not binance_success:
        await update.message.reply_text("⚠️ Не удалось получить данные о курсах")
        return

    message = []
    time_updates = []

    # Блок ЦБ РФ
    if cbr_success :
        message.append("🏛️ <b>Официальные курсы от ЦБ РФ</b>\n\n")
        currencies_cbr = {
            'USD' : {'symbol' : '💵', 'name' : 'Доллар США'},
            'EUR' : {'symbol' : '💶', 'name' : 'Евро'},
            'CNY' : {'symbol' : '💴', 'name' : 'Китайский юань'}
        }

        for code, info in currencies_cbr.items() :
            data = currency_api_cbr.get_currency_info(code)
            if not data :
                continue

            change_abs = data['change_abs']
            change_pct = data['change_pct']

            # Определение стрелки
            if change_pct > 0 :
                arrow = "↗️"
            elif change_pct < 0 :
                arrow = "↘️"
            else :
                arrow = "➖"

            # Форматирование изменения
            if change_pct == 0 and change_abs == 0 :
                change_line = "➖"
            else :
                change_abs_str = f"{change_abs:+.2f}"
                pct_str = f"{abs(change_pct):.2f}%"
                change_line = f"{change_abs_str} ({pct_str} {arrow})"

            message.append(
                f"{info['symbol']} <b>{info['name']}</b>\n"
                f"➤ Текущий: {data['current']} ₽\n"
                f"➤ Изменение: {change_line}\n\n"
            )
        time_updates.append(f"ЦБ РФ: {currency_api_cbr.last_update.strftime('%d.%m.%Y %H:%M')}")
        message.append("______________________________\n")

    # Блок ExchangeRate-API
    if exchange_success :
        message.append("\n🌍 <b>Независимые курсы от ExchangeRate</b>\n\n")  # Добавлен \n

        # Доллар (прямой курс)
        usd_rate = round(exchange_api.usd_rate, 2)
        message.append(
            f"💵 Доллар США\n"
            f"➤ Курс: {usd_rate} ₽\n\n"
        )

        # Евро (через конвертацию)
        eur_rate = exchange_api.convert_via_usd('EUR')
        if eur_rate :
            message.append(
                f"💶 Евро\n"
                f"➤ Курс: {eur_rate} ₽\n\n"
            )

        # Юань (через конвертацию)
        cny_rate = exchange_api.convert_via_usd('CNY')
        if cny_rate :
            message.append(
                f"💴 Китайский юань\n"
                f"➤ Курс: {cny_rate} ₽\n\n"
            )
        time_updates.append(f"ExchangeRate: {exchange_api.last_update.strftime('%d.%m.%Y %H:%M')}")
        message.append("______________________________\n")

    # Блок Binance
    if binance_success :
        message.append("\n🌐 <b>Курсы криптовалют от Binance</b>\n\n")
        binance_times = []

        # Спотовые курсы
        for pair_name, pair_data in binance_spot.items() :
            if pair_data :
                binance_times.append(pair_data['time'])
                currency_symbol = '₽' if 'RUB' in pair_name else '$'
                change_abs = pair_data['change_abs']
                change_pct = pair_data['change_pct']

                # Определение стрелки
                if change_pct > 0 :
                    arrow = "↗️"
                elif change_pct < 0 :
                    arrow = "↘️"
                else :
                    arrow = "➖"

                # Форматирование изменения
                if change_pct == 0 and change_abs == 0 :
                    change_str = "➖"
                else :
                    change_abs_str = f"{change_abs:+.2f}"
                    pct_str = f"{abs(change_pct):.2f}%"
                    change_str = f"{change_abs_str} ({pct_str} {arrow})"

                message.append(
                    f"<b>Пара {pair_name.replace('_', '/')}</b>\n"
                    f"➤ Текущий курс: {currency_symbol}{pair_data['price']:,.2f}\n"
                    f"➤ Изменение: {change_str}\n\n"
                )

        # P2P курсы
        if binance_p2p_buy and binance_p2p_sell :
            message.append(
                f"<b>USDT/RUB (P2P)</b>\n"
                f"➤ Покупка: ₽{binance_p2p_buy['price']:,.2f}\n"
                f"➤ Продажа: ₽{binance_p2p_sell['price']:,.2f}\n\n"
            )
            binance_times.extend([binance_p2p_buy['time'], binance_p2p_sell['time']])

        # Время обновления
        if binance_times :
            latest_time = max(binance_times)
            time_updates.append(f"Binance: {latest_time.strftime('%d.%m.%Y %H:%M')}")

        message.append("______________________________\n")

    # Форматирование времени обновления
    if time_updates :
        message.append("\n🕒 <b>Время обновления:</b>")
        ordered_updates = []
        for service in ["ЦБ РФ", "ExchangeRate", "Binance"] :
            for update_time in time_updates :
                if update_time.startswith(service) :
                    ordered_updates.append(update_time)
                    break
        message.extend([f"\n{time_str}" for time_str in ordered_updates])

    # Отправка сообщения
    await update.message.reply_text(
        text="\n".join(message).replace('\n\n', '\n'),
        parse_mode='HTML'
    )

async def catalog_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # Создаем кнопку с web_app, указываем URL мини-приложения
    web_app_info = WebAppInfo(url="https://bestpriceweb.ru")  # <-- ваш HTTPS-домен
    button = InlineKeyboardButton("Открыть каталог", web_app=web_app_info)
    keyboard = [[button]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Отправляем сообщение с Inline-кнопкой
    await update.message.reply_text(
        text="Нажмите кнопку, чтобы открыть каталог",
        reply_markup=reply_markup
    )
    pass




# Регистрируем этот хендлер ДО остальных



# ------------------------------ MAIN ------------------------------
async def main() -> None:
    # ── Инициализируем БД ───────────────────────────────────────────────
    init_db()

    # ── Создаём приложение Telegram-бота ────────────────────────────────
    application = ApplicationBuilder().token(API_TOKEN).build()

    # ── Меню / команды бота ─────────────────────────────────────────────
    commands = [
        BotCommand("start",         "🏁 Запустить бота"),
        BotCommand("best",          "🔥 Лучшие предложения"),
        BotCommand("markup",        "💰 Наценка на товар"),
        BotCommand("gradmark",      "📈 Наценка с градацией"),
        BotCommand("list",          "🔍 Сравнение цен"),
        BotCommand("my_price_list", "📊 Свой прайс-лист"),
        BotCommand("clear",         "♻️ Очистить список товаров"),
        BotCommand("currency",      "💲️ Курсы валют"),
        BotCommand("help",          "⚙️ Получить справку"),
        BotCommand("restart",       "🛠️ Сброс всех состояний"),
        # BotCommand("ai_assistant",  "🤖 AI-форматирование прайса (beta)"), ⚒️
        # BotCommand("catalog", "📚 Каталог товаров"), ⚒️
    ]
    await application.bot.set_my_commands(commands)

    # ── Обычные команды ────────────────────────────────────────────────
    application.add_handler(CommandHandler("start",  start))
    application.add_handler(CommandHandler("help",   help_command))
    application.add_handler(CommandHandler("best",   best_command))
    application.add_handler(CommandHandler("restart_price_list", restart_price_list))
    application.add_handler(CommandHandler("clear",  clear_command))
    application.add_handler(CommandHandler("currency", currency_command))
    application.add_handler(CommandHandler("restart", restart_all), group=0)
    # application.add_handler(CommandHandler("catalog", catalog_command)) ⚒️

    # ── AI-ассистент (ConversationHandler) ─────────────────────────────
    # register_ai_assistant(application) ⚒️

    # ── ConversationHandler’ы основных функций ────────────────────────
    application.add_handler(get_my_price_list_conversation_handler())
    application.add_handler(get_list_conversation_handler())
    application.add_handler(get_markup_conversation_handler())
    application.add_handler(get_gradmark_conversation_handler())

    # ── Callback-хендлеры для /best ────────────────────────────────────
    application.add_handler(CallbackQueryHandler(best_group_callback,      pattern=r"^group_"))
    application.add_handler(CallbackQueryHandler(best_category_callback,   pattern=r"^cat_"))
    application.add_handler(CallbackQueryHandler(best_command_callback,    pattern=r"^best_(csv|excel|msg)$"))
    application.add_handler(CallbackQueryHandler(best_command_comments_callback,
                                                 pattern=r"^best_msg_comments_"))
    application.add_handler(CallbackQueryHandler(best_structure_callback,
                                                 pattern=r"^(structure_grouped|structure_flat)$"))

    # ── Общий обработчик сообщений ─────────────────────────────────────
    application.add_handler(MessageHandler(filters.ALL, message_handler))

    print("Бот запущен. Нажмите Ctrl+C для остановки.")
    await application.run_polling()


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
